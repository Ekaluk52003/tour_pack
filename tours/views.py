# tour_quote/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from .models import TourPackageQuote, City, Hotel, Service, GuideService, ServiceType, TourDay, TourDayService, TourDayGuideService, PredefinedTourQuote, ReferenceID, ServicePrice, TourPackType, Agency, Invoice, InvoiceItem, SupplierExpense, InvoiceReferenceID, Supplier, SupplierService, ServiceExpenseTemplate
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField, Case, When, IntegerField
from django.db.models.functions import Coalesce
import json
from django.views.decorators.http import require_http_methods
from django.urls import reverse
from django.core.serializers.json import DjangoJSONEncoder
from weasyprint import HTML, CSS
from django.template.loader import render_to_string
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Prefetch, Q
import traceback
from decimal import Decimal, InvalidOperation
import json
import logging
from django.conf import settings
from django.core.mail import EmailMessage
from io import BytesIO
from django.contrib import messages
import os
import base64
from datetime import datetime, timedelta
import uuid
import csv
import io

logger = logging.getLogger(__name__)


def parse_custom_date(value):
    """Parse a date string in various formats (e.g., '25-Dec-26', '2026-12-25') to a date object."""
    if not value or value.strip() == '':
        return None
    value = value.strip()
    formats = ['%d-%b-%y', '%d-%b-%Y', '%Y-%m-%d']
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def safe_decimal(value, default=Decimal('0')):
    if isinstance(value, (list, tuple)):
        return default  # Return default if value is a sequence
    try:
        return Decimal(str(value))
    except (ValueError, TypeError, InvalidOperation):
        logger.warning(
            f"Failed to convert {value} to Decimal. Using default value {default}")
        return default


def safe_float(value, default=0.0):
    """Safely convert a value to float, returning default if conversion fails or value is empty."""
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def calculate_totals(package):
    service_total = sum(
        service.price_at_booking for day in package.tour_days.all() for service in day.services.all()
    )

    guide_service_total = sum(
        guide_service.price_at_booking for day in package.tour_days.all() for guide_service in day.guide_services.all()
    )

    hotel_total = sum(
        (Decimal(cost['price']) * int(cost['room']) * int(cost['nights'])) +
        (Decimal(cost.get('extraBedPrice', 0)) * int(cost['nights']))
        for cost in package.hotel_costs
    )

    service_grand_total = Decimal(service_total) + guide_service_total
    hotel_grand_total = Decimal(hotel_total)
    total_discount = sum(Decimal(discount['amount'])
                         for discount in package.discounts)
    total_extra_cost = sum(Decimal(extra_cost['amount']) for extra_cost in package.extra_costs)
    grand_total = Decimal(service_total) + guide_service_total + hotel_total + total_extra_cost - total_discount

    return service_grand_total, hotel_grand_total, grand_total, total_discount, total_extra_cost


@login_required
@require_http_methods(["POST"])
def save_tour_package(request, package_reference=None):
    try:
        data = json.loads(request.body)
        with transaction.atomic():
            # Determine if it's a new package or an update
            if package_reference:
                package = get_object_or_404(TourPackageQuote, package_reference=package_reference)
                is_new = False
            else:
                # Check if user can create packages (superuser or assistance group)
                can_create = request.user.is_superuser or request.user.groups.filter(name='assistance').exists()
                if can_create:
                    package = TourPackageQuote()
                    package.prepare_by_user = request.user
                    is_new = True
                else:
                    return JsonResponse({'status': 'error', 'message': 'You do not have permission to create new packages'}, status=403)

            # Check if user can edit basic fields (superuser or assistance group)
            can_edit = request.user.is_superuser or request.user.groups.filter(name='assistance').exists()
            if can_edit:
                # Users with edit permissions can edit basic fields
                package.name = data['name']
                package.customer_name = data['customer_name']
                package.billing_name = data.get('billing_name', '')
                package.remark = data.get('remark', '')
                package.connection_ref = data.get('connectionRef', '')
                package.remark2 = data.get('remark2', '')
                package.tour_pack_type_id = data['tour_pack_type']

                # Only superusers can edit commission rates
                if request.user.is_superuser:
                    package.commission_rate_hotel = data.get(
                        'commission_rate_hotel', 0)
                    package.commission_rate_services = data.get(
                        'commission_rate_services', 0)
                # For assistance group, set default commission rates only on creation
                elif request.user.groups.filter(name='assistance').exists() and is_new:
                    package.commission_rate_hotel = 20
                    package.commission_rate_services = 5

                # Note: prepare_by_user is only set during creation, never during updates
                # This field is immutable after the tour quote is created


            # all login users can update these field
            package.hotel_costs = data['hotelCosts']
            package.alternative_hotels = data.get('alternativeHotels', [])
            package.remark_of_hotels = data.get('remark_of_hotels', '')
            package.special_note = data.get('special_note', '')
            package.billing_name = data.get('billing_name', '')
            package.connection_ref = data.get('connectionRef', '')
            package.discounts = data.get('discounts', [])
            package.extra_costs = data.get('extraCosts', [])

            # Save the package to get a primary key
            package.save()

            if can_edit or is_new:
                # Clear existing days and services if it's an update
                if not is_new:
                    package.tour_days.all().delete()

                # Create new days and services
                for day_data in data['days']:
                    tour_day = TourDay.objects.create(
                        tour_package=package,
                        date=day_data['date'],
                        city_id=day_data['city'],
                        hotel_id=day_data['hotel']
                    )

                    # Create services for the day
                    for service_data in day_data['services']:
                        service = Service.objects.get(id=service_data['name'])
                        TourDayService.objects.create(
                            tour_day=tour_day,
                            service=service,
                            price_at_booking=service_data['price_at_booking']
                        )

                    # Create guide services for the day
                    for guide_service_data in day_data['guide_services']:
                        guide_service = GuideService.objects.get(id=guide_service_data['name'])
                        TourDayGuideService.objects.create(
                            tour_day=tour_day,
                            guide_service=guide_service,
                            price_at_booking=guide_service_data['price_at_booking']
                        )

            # Recalculate totals
            service_grand_total, hotel_grand_total, grand_total, total_discount, total_extra_cost = calculate_totals(package)

            package.service_grand_total = service_grand_total
            package.hotel_grand_total = hotel_grand_total
            package.grand_total_cost = grand_total

            # Recalculate commission amounts
            total_room_nights = sum(
                safe_decimal(hotel.get('room')) *
                safe_decimal(hotel.get('nights'))
                for hotel in package.hotel_costs
            )

            package.commission_rate_hotel = safe_decimal(
                package.commission_rate_hotel)
            package.commission_rate_services = safe_decimal(
                package.commission_rate_services)

            package.commission_amount_hotel = package.commission_rate_hotel * total_room_nights
            package.commission_amount_services = package.commission_rate_services * \
                package.service_grand_total / Decimal('100')

            package.save()  # Save the package with all updates

        return JsonResponse({
            'status': 'success',
            'message': 'Tour package saved successfully',
            'package_reference': package.package_reference
        })


    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data provided'}, status=400)
    except TourPackageQuote.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Tour package not found'}, status=404)
    except PermissionError:
        return JsonResponse({'status': 'error', 'message': 'You do not have permission to perform this action'}, status=403)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'An unexpected error occurred: {str(e)}'}, status=500)


def _build_alternative_hotels_with_total(alternative_hotels):
    result = []
    for alt in alternative_hotels:
        try:
            room_cost = float(alt.get('room') or 0) * float(alt.get('nights') or 0) * float(alt.get('price') or 0)
        except (ValueError, TypeError):
            room_cost = 0
        extra_bed_price = alt.get('extraBedPrice', '')
        try:
            extra_bed_cost = float(extra_bed_price) * float(alt.get('nights') or 0) if extra_bed_price and str(extra_bed_price).strip() else 0
        except (ValueError, TypeError):
            extra_bed_cost = 0
        alt_with_total = alt.copy()
        alt_with_total['total'] = room_cost + extra_bed_cost
        result.append(alt_with_total)
    return result


@login_required
def tour_package_pdf(request, pk):
    package = get_object_or_404(TourPackageQuote, pk=pk)

    # Define the image path based on your project structure
    logo_data_uri = None

    # Try multiple possible paths for Docker environment
    possible_paths = [
      os.path.join(settings.BASE_DIR, 'static', 'image', 'rsz_animo1.png')        # Add more paths if needed
    ]

    # Try to find and load the image
    for logo_path in possible_paths:

        if os.path.exists(logo_path):
            try:
                with open(logo_path, 'rb') as f:
                    logo_binary = f.read()
                    logo_base64 = base64.b64encode(logo_binary).decode('utf-8')
                    logo_data_uri = f'data:image/png;base64,{logo_base64}'
                    break
            except Exception as e:
                print(f"Error reading file at {logo_path}: {str(e)}")  # Debug print

    if not logo_data_uri:
        print("Could not find or load the logo image")  # Debug print

    # Your existing code for hotel costs calculations...
    hotel_costs_with_total = []
    for cost in package.hotel_costs:
        room_cost = float(cost['room']) * \
            float(cost['nights']) * float(cost['price'])
        extra_bed_price = cost.get('extraBedPrice', '')
        extra_bed_cost = float(extra_bed_price) * float(
            cost['nights']) if extra_bed_price and extra_bed_price.strip() else 0
        total_cost = room_cost + extra_bed_cost

        cost_with_total = cost.copy()
        cost_with_total['room_cost'] = room_cost
        cost_with_total['extra_bed_cost'] = extra_bed_cost
        cost_with_total['total'] = total_cost

        hotel_costs_with_total.append(cost_with_total)

    discounts = package.discounts
    total_discount = sum(float(discount['amount']) for discount in discounts)

    extra_costs = package.extra_costs
    total_extra_cost = sum(float(extra_cost['amount']) for extra_cost in extra_costs)

    remark2 = package.remark2.replace(
        '\n', '<br>') if package.remark2 is not None else ''

    remark_of_hotels = package.remark_of_hotels.replace(
        '\n', '<br>') if package.remark_of_hotels is not None else ''

    special_note = package.special_note.replace(
        '\n', '<br>') if package.special_note is not None else ''

    ordered_tour_days = package.tour_days.all().order_by('date')

    html_string = render_to_string('tour_quote/tour_package_pdf.html', {
        'package': package,
        'ordered_tour_days': ordered_tour_days,
        'tour_pack_type': package.tour_pack_type,
        'hotel_costs_with_total': hotel_costs_with_total,
        'base_url': request.build_absolute_uri('/'),
        'discounts': discounts,
        'total_discount': total_discount,
        'extra_costs': extra_costs,
        'total_extra_cost': total_extra_cost,
        'static_url': settings.STATIC_URL,
        'remark2': remark2,
        'remark_of_hotels': remark_of_hotels,
        'special_note': special_note,
        'logo_data_uri': logo_data_uri,
        'alternative_hotels': _build_alternative_hotels_with_total(package.alternative_hotels),
    })

    # Create response and set filename
    response = HttpResponse(content_type='application/pdf')
    package_name = package.name if package.name else 'unknown'
    reference = package.package_reference if package.package_reference else str(package.id)
    package_name = ''.join(e for e in package_name if e.isalnum() or e in ['-', '_','(', ')']).strip()
    reference = ''.join(e for e in reference if e.isalnum() or e in ['-', '_']).strip()
    filename = f"{package_name}_{reference}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    # Generate PDF with custom styles
    HTML(
        string=html_string,
        base_url=request.build_absolute_uri()
    ).write_pdf(
        response,
        stylesheets=[CSS(string='''
            @page {
                size: A4;
                margin: 2cm;
                @bottom-right {
                    content: "Page " counter(page) " of " counter(pages);
                    font-size: 10px;
                    color: #666;
                }
            }
            body {
                font-family: sans-serif;
            }
        ''')]
    )

    return response


@login_required
def tour_package_pdf_no_cost(request, pk):
    package = get_object_or_404(TourPackageQuote, pk=pk)

    # Define the image path based on your project structure
    logo_data_uri = None

    # Try multiple possible paths for Docker environment
    possible_paths = [
      os.path.join(settings.BASE_DIR, 'static', 'image', 'rsz_animo1.png')        # Add more paths if needed
    ]

    # Try to find and load the image
    for logo_path in possible_paths:
        if os.path.exists(logo_path):
            try:
                with open(logo_path, 'rb') as f:
                    logo_binary = f.read()
                    logo_base64 = base64.b64encode(logo_binary).decode('utf-8')
                    logo_data_uri = f'data:image/png;base64,{logo_base64}'
                    break
            except Exception as e:
                print(f"Error reading file at {logo_path}: {str(e)}")  # Debug print

    if not logo_data_uri:
        print("Could not find or load the logo image")  # Debug print

    # We still need the hotel costs structure but without prices for the template
    hotel_costs_with_total = []
    for cost in package.hotel_costs:
        cost_with_total = cost.copy()
        # Keep structure but zero out costs
        cost_with_total['room_cost'] = 0
        cost_with_total['extra_bed_cost'] = 0
        cost_with_total['total'] = 0
        hotel_costs_with_total.append(cost_with_total)

    # We keep the structure but zero out all costs
    discounts = package.discounts
    extra_costs = package.extra_costs

    remark2 = package.remark2.replace(
        '\n', '<br>') if package.remark2 is not None else ''
    remark_of_hotels = package.remark_of_hotels.replace(
        '\n', '<br>') if package.remark_of_hotels is not None else ''

    special_note = package.special_note.replace(
        '\n', '<br>') if package.special_note is not None else ''

    ordered_tour_days = package.tour_days.all().order_by('date')

    html_string = render_to_string('tour_quote/tour_package_pdf_no_cost.html', {
        'package': package,
        'ordered_tour_days': ordered_tour_days,
        'tour_pack_type': package.tour_pack_type,
        'hotel_costs_with_total': hotel_costs_with_total,
        'base_url': request.build_absolute_uri('/'),
        'discounts': discounts,
        'extra_costs': extra_costs,
        'static_url': settings.STATIC_URL,
        'remark2': remark2,
        'remark_of_hotels': remark_of_hotels,
        'special_note': special_note,
        'logo_data_uri': logo_data_uri,
        'hide_costs': True,  # Flag to hide costs in template
        'alternative_hotels': package.alternative_hotels,
    })

    # Create response and set filename
    response = HttpResponse(content_type='application/pdf')
    package_name = package.name if package.name else 'unknown'
    reference = package.package_reference if package.package_reference else str(package.id)
    package_name = ''.join(e for e in package_name if e.isalnum() or e in ['-', '_','(', ')']).strip()
    reference = ''.join(e for e in reference if e.isalnum() or e in ['-', '_']).strip()
    filename = f"{package_name}_{reference}_no_cost.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    # Generate PDF with custom styles
    HTML(
        string=html_string,
        base_url=request.build_absolute_uri()
    ).write_pdf(
        response,
        stylesheets=[CSS(string='''
            @page {
                size: A4;
                margin: 2cm;
                @bottom-right {
                    content: "Page " counter(page) " of " counter(pages);
                    font-size: 10px;
                    color: #666;
                }
            }
            body {
                font-family: sans-serif;
            }
        ''')]
    )

    return response


@login_required
@require_http_methods(["POST"])
def send_tour_package_email(request, pk):
    try:
        package = get_object_or_404(TourPackageQuote, pk=pk)

        # Define the image path based on your project structure
        logo_data_uri = None
        # Try multiple possible paths for Docker environment
        possible_paths = [
        os.path.join(settings.BASE_DIR, 'static', 'image', 'rsz_animo1.png')         ]

        # Try to find and load the image
        for logo_path in possible_paths:
            if os.path.exists(logo_path):
                try:
                    with open(logo_path, 'rb') as f:
                        logo_binary = f.read()
                        logo_base64 = base64.b64encode(logo_binary).decode('utf-8')
                        logo_data_uri = f'data:image/png;base64,{logo_base64}'
                        break
                except Exception as e:
                    print(f"Error reading file at {logo_path}: {str(e)}")  # Debug print

        if not logo_data_uri:
            print("Could not find or load the logo image")  # Debug print
        cc_email = request.POST.get('cc_email')


        # Calculate totals (replicating logic from tour_package_pdf)
        hotel_costs_with_total = []
        for cost in package.hotel_costs:
            room_cost = float(cost['room']) * \
                float(cost['nights']) * float(cost['price'])
            extra_bed_price = cost.get('extraBedPrice', '')
            extra_bed_cost = float(extra_bed_price) * float(
                cost['nights']) if extra_bed_price and extra_bed_price.strip() else 0
            total_cost = room_cost + extra_bed_cost

            cost_with_total = cost.copy()
            cost_with_total['room_cost'] = room_cost
            cost_with_total['extra_bed_cost'] = extra_bed_cost
            cost_with_total['total'] = total_cost

            hotel_costs_with_total.append(cost_with_total)

        discounts = package.discounts
        total_discount = sum(float(discount['amount'])
                             for discount in discounts)

        extra_costs = package.extra_costs
        total_extra_cost = sum(float(extra_cost['amount']) for extra_cost in extra_costs)

        remark2 = package.remark2.replace(
        '\n', '<br>') if package.remark2 is not None else ''

        remark_of_hotels = package.remark_of_hotels.replace(
        '\n', '<br>') if package.remark_of_hotels is not None else ''

        # Get ordered tour days
        ordered_tour_days = package.tour_days.all().order_by('date')
        # Render the template to HTML
        html_string = render_to_string('tour_quote/tour_package_pdf.html', {
            'package': package,
            'ordered_tour_days': ordered_tour_days,
            'tour_pack_type': package.tour_pack_type,
            'hotel_costs_with_total': hotel_costs_with_total,
            'base_url': request.build_absolute_uri('/'),
            'discounts': discounts,
            'total_discount': total_discount,
            'extra_costs': extra_costs,
            'total_extra_cost': total_extra_cost,
            'static_url': settings.STATIC_URL,
            'remark2': remark2,
            'remark_of_hotels':remark_of_hotels,
            'logo_data_uri': logo_data_uri
        })

        # Generate PDF
        pdf_file = BytesIO()
        HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(pdf_file, stylesheets=[CSS(string='''
            @page {
                size: A4;
                margin: 2cm;
                @bottom-right {
                    content: "Page " counter(page) " of " counter(pages);
                    font-size: 10px;
                    color: #666;
                }
            }
            body {
                font-family: sans-serif;
            }
        ''')])
        pdf_file.seek(0)


          # Generate filename
        package_name = package.name if package.name else 'unknown'
        reference = package.package_reference if package.package_reference else str(package.id)

        # Clean the filename components
        package_name = ''.join(e for e in package_name if e.isalnum() or e in ['-', '_','(', ')']).strip()
        reference = ''.join(e for e in reference if e.isalnum() or e in ['-', '_']).strip()

        filename = f"{package_name}_{reference}.pdf"

        # Prepare email
        subject = f'Tour Package {package.package_reference}: {package.name}'
        message = f'Please find attached the tour package quote for {package.customer_name}.'
        from_email = settings.DEFAULT_FROM_EMAIL
        # You might want to change this to use the customer's email
        to_email = ['jimforanimo@gmail.com']
        if cc_email:
            to_email.append(cc_email)

        # Send email
        email = EmailMessage(subject, message, from_email, to_email)
        email.attach(filename, pdf_file.getvalue(), 'application/pdf')

        email.send(fail_silently=False)

        context = {
            'message': f'Email sent successfully to {", ".join(to_email)}',
            'class': 'text-green-600'
        }
        return render(request, 'tour_quote/notification.html', context)
    except Exception as e:
        logger.error(f"Error in send_tour_package_email: {str(e)}")
        context = {
            'message': f'Failed to send email: {str(e)}',
            'class': 'text-red-600'
        }
        return render(request, 'tour_quote/notification.html', context)




@login_required
def tour_package_detail(request, package_reference):
    package = get_object_or_404(TourPackageQuote, package_reference=package_reference)

    # Calculate total costs for hotels
    hotel_costs_with_total = []
    for cost in package.hotel_costs:
        room_cost = float(cost['room']) * \
            float(cost['nights']) * float(cost['price'])

        extra_bed_price = cost.get('extraBedPrice', '')
        extra_bed_cost = float(extra_bed_price) * float(
            cost['nights']) if extra_bed_price and extra_bed_price.strip() else 0

        total_cost = room_cost + extra_bed_cost

        cost_with_total = cost.copy()  # Create a copy to avoid modifying the original
        cost_with_total['room_cost'] = room_cost
        cost_with_total['extra_bed_cost'] = extra_bed_cost
        cost_with_total['total'] = total_cost

        hotel_costs_with_total.append(cost_with_total)

      # Prepare discount information

    discounts = package.discounts
    total_discount = sum(float(discount['amount']) for discount in discounts)

    # Prepare extra costs information
    extra_costs = package.extra_costs
    total_extra_cost = sum(float(extra_cost['amount']) for extra_cost in extra_costs)
    ordered_tour_days = package.tour_days.all().order_by('date')
    remark2 = package.remark2.replace(
        '\n', '<br>') if package.remark2 is not None else ''
    # remark2 = package.remark2.replace('\n', '<br>')

    billing_name = package.billing_name.replace(
        '\n', '<br>') if package.billing_name is not None else ''

    comission_total = package.commission_amount_hotel + package.commission_amount_services
    # Check if user should see commission info (not in assistance group)
    show_commission = not request.user.groups.filter(name='assistance').exists()

    tour_day_hotel_names = {day.hotel.name for day in ordered_tour_days if day.hotel}
    hotel_cost_names = {cost.get('name', '') for cost in package.hotel_costs if cost.get('name')}
    hotel_mismatch_missing = sorted(tour_day_hotel_names - hotel_cost_names)
    hotel_mismatch_orphaned = sorted(hotel_cost_names - tour_day_hotel_names)

    context = {
        'package': package,
        # Pass hotel costs with total calculation
        'hotel_costs_with_total': hotel_costs_with_total,
        'alternative_hotels': _build_alternative_hotels_with_total(package.alternative_hotels),
        'hotel_mismatch_missing': hotel_mismatch_missing,
        'hotel_mismatch_orphaned': hotel_mismatch_orphaned,
        'tour_pack_type': package.tour_pack_type,  # Add this line
        'discounts': discounts,
        'total_discount': total_discount,
        'extra_costs': extra_costs,
        'total_extra_cost': total_extra_cost,
        'remark2': remark2,
        'billing_name': billing_name,
        'comission_total': comission_total,
        'ordered_tour_days': ordered_tour_days,
        'show_commission': show_commission,
    }

    return render(request, 'tour_quote/tour_package_detail.html', context)


@login_required
def tour_package_edit(request, package_reference):
    package = get_object_or_404(TourPackageQuote, package_reference=package_reference)
    cities = City.objects.all()
    guide_services = list(GuideService.objects.values('id', 'name', 'price'))
    all_hotels = list(Hotel.objects.values('id', 'name', 'city__name'))
    predefined_quotes = PredefinedTourQuote.objects.all()
    tour_pack_types = TourPackType.objects.all()

    if request.method == 'POST':
        return save_tour_package(request, package.id)

    package_data = {
        'package_reference': package.package_reference,
        'name': package.name,
        'customer_name': package.customer_name,
        'billing_name': package.billing_name,
        'remark': package.remark,
        'connectionRef': package.connection_ref,
        'remark2': package.remark2,
        'remark_of_hotels': package.remark_of_hotels,
        'special_note': package.special_note,
        'tour_pack_type': package.tour_pack_type_id,
        # Add hotel commission rate
        'commission_rate_hotel': float(package.commission_rate_hotel),
        'commission_rate_services': float(package.commission_rate_services),
        'days': [
            {
                'date': day.date.isoformat(),
                'city': str(day.city_id),
                'hotel': str(day.hotel_id),
                'city_name': day.city.name,
                'hotel_name': day.hotel.name,
                'services': [
                    {
                        'type': service.service.service_type.name.lower(),
                        'name': str(service.service_id),
                        'service_name': service.service.name,
                        'price_at_booking': float(service.price_at_booking)
                    }
                    for service in day.services.all()
                ],
                'guideServices': [
                    {
                        'name': str(guide_service.guide_service_id),
                        'price': float(guide_service.guide_service.price),
                        'price_at_booking': float(guide_service.price_at_booking)
                    }
                    for guide_service in day.guide_services.all()
                ]
            }
            for day in package.tour_days.all().order_by('date')
        ],
        'discounts': [
            {
                'item': discount['item'],
                'amount': float(discount['amount'])
            }
            for discount in package.discounts
        ],

          'extraCosts': [
            {
                'item': extra_cost['item'],
                'amount': safe_float(extra_cost.get('amount', 0)),
                'price': safe_float(extra_cost.get('price', 0)),
                'qty': safe_float(extra_cost.get('qty', 1), 1.0)
            }
            for extra_cost in package.extra_costs
        ],
        'hotelCosts': package.hotel_costs,
        'alternativeHotels': package.alternative_hotels,

    }

    # Check if user should see commission info (not in assistance group)
    show_commission = not request.user.groups.filter(name='assistance').exists()
    # Check if user can edit (superuser or assistance group)
    can_edit = request.user.is_superuser or request.user.groups.filter(name='assistance').exists()

    context = {
        'package': package,
        'cities': cities,
        'guide_services_json': json.dumps(guide_services, cls=DjangoJSONEncoder).replace('</', '<\\/'),
        'all_hotels_json': json.dumps(all_hotels, cls=DjangoJSONEncoder).replace('</', '<\\/'),
        'package_json': json.dumps(package_data, cls=DjangoJSONEncoder).replace('</', '<\\/'),
        'predefined_quotes': predefined_quotes,
        'tour_pack_types': tour_pack_types,
        'show_commission': show_commission,
        'can_edit': can_edit,
    }

    return render(request, 'tour_quote/tour_package_edit.html', context)


@login_required
def tour_package_quote(request):
    cities = City.objects.all()
    service_types = ServiceType.objects.all()

    predefined_quotes = PredefinedTourQuote.objects.all()
    tour_pack_types = TourPackType.objects.all()
    # Query guide services and convert price (Decimal) to float
    guide_services = list(
        GuideService.objects.values('id', 'name', 'price')
    )
    # Convert the Decimal to float
    for guide_service in guide_services:
        guide_service['price'] = float(guide_service['price'])

    print("Guide Services:", guide_services)

    # Check if user can use AI email parser
    # can_use_ai = request.user.is_superuser or request.user.groups.filter(name='assistance').exists()

    # only superuser can use AI email parser
    can_use_ai = request.user.is_superuser

    context = {
        'cities': cities,
        'service_types': service_types,
        'guide_services_json': json.dumps(guide_services, cls=DjangoJSONEncoder),
        'predefined_quotes': predefined_quotes,
        'tour_pack_types': tour_pack_types,
        'is_superuser': request.user.is_superuser,
        'can_use_ai': can_use_ai,
    }

    return render(request, 'tour_quote/tour_package_quote.html', context)


@login_required
@require_http_methods(["GET"])
def get_city_services(request, city_id):
    try:
        tour_pack_type_id = request.GET.get('tour_pack_type')
        city = get_object_or_404(City, id=city_id)
        tour_pack_type = get_object_or_404(TourPackType, id=tour_pack_type_id)

        hotels = list(Hotel.objects.filter(city=city).values('id', 'name'))

        services = Service.objects.filter(
            cities=city
        ).prefetch_related('prices').select_related('service_type')

        service_types = {}
        for service in services:
            service_type = service.service_type.name
            if service_type not in service_types:
                service_types[service_type] = []

            price = service.prices.filter(tour_pack_type=tour_pack_type).first()
            if price:
                service_types[service_type].append({
                    'id': service.id,
                    'name': service.name,
                    'price': float(price.price)
                })

        guide_services = list(
            GuideService.objects.all().values('id', 'name', 'price'))
        for gs in guide_services:
            gs['price'] = float(gs['price'])

        response_data = {
            'hotels': hotels,
            'service_types': [
                {'type': st, 'services': services}
                for st, services in service_types.items()
            ],
            'guide_services': guide_services
        }

        return JsonResponse(response_data, safe=False, content_type='application/json')
    except Exception as e:
        print("Error in get_city_services:", str(e))
        return JsonResponse({'error': str(e)}, status=500, content_type='application/json')


@login_required
def tour_packages(request):
    query = request.GET.get('q', '')  # Get the search query

    # Filter packages based on search query
    if query:
        packages = TourPackageQuote.objects.filter(name__icontains=query) | TourPackageQuote.objects.filter(
            customer_name__icontains=query) | TourPackageQuote.objects.filter(package_reference__icontains=query) | TourPackageQuote.objects.filter(
            prepare_by_user__username__icontains=query) | TourPackageQuote.objects.filter(connection_ref__icontains=query)
    else:
        packages = TourPackageQuote.objects.all()

    # Pagination: Show 10 packages per page

    packages = packages.order_by('-created_at')

    paginator = Paginator(packages, 10)
    page_number = request.GET.get('page')
    packages_page = paginator.get_page(page_number)

    context = {
        'packages': packages_page,
        'query': query,
    }
    return render(request, 'tour_quote/tour_packages.html', context)


@login_required
@require_http_methods(["GET"])
def get_predefined_tour_quote(request, quote_id):
    try:
        quote = get_object_or_404(PredefinedTourQuote.objects, id=quote_id)
        tour_pack_type_id = request.GET.get('tour_pack_type')

        if not tour_pack_type_id:
            return JsonResponse({'error': 'Tour package type is required'}, status=400)

        tour_pack_type = get_object_or_404(TourPackType, id=tour_pack_type_id)
        days = []

        for day in quote.days.all().select_related('city', 'hotel'):
            day_data = {
                'city': day.city.id,
                'hotel': day.hotel.id,
                'services': [],
                'guideServices': []
            }

            # Get services for the day
            for day_service in day.services.all().select_related('service__service_type'):
                service = day_service.service

                # Check if this service is available in this city
                if service.cities.filter(id=day.city.id).exists():
                    # Get price for the selected tour pack type
                    try:
                        price = ServicePrice.objects.get(
                            service=service,
                            tour_pack_type=tour_pack_type
                        ).price
                    except ServicePrice.DoesNotExist:
                        price = None

                    if price is not None:
                        day_data['services'].append({
                            'id': service.id,
                            'name': service.name,
                            'type': service.service_type.name,
                            'price': float(price),
                            'quantity': day_service.quantity,
                            'order': day_service.order
                        })

            # Get guide services
            for guide_service in day.guide_services.all().select_related('guide_service'):
                day_data['guideServices'].append({
                    'id': guide_service.guide_service.id,
                    'name': guide_service.guide_service.name,
                    'price': float(guide_service.guide_service.price)
                })

            days.append(day_data)

        response_data = {
            'id': quote.id,
            'name': quote.name,
            'description': quote.description,
            'days': days
        }

        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"Error in get_predefined_tour_quote: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)




@login_required
@require_http_methods(["GET"])
def duplicate_tour_package(request, package_reference):
    # Check if the user is a superuser
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to duplicate tour packages.")
        return redirect('tour_package_detail', package_reference=package_reference)

    # Get the original package
    original_package = get_object_or_404(TourPackageQuote, package_reference=package_reference)

    # Create a new package with copied data
    new_package = TourPackageQuote.objects.create(
        name=f"{original_package.name}_(copy)",
        customer_name=original_package.customer_name,
        remark=original_package.remark,
        connection_ref=original_package.connection_ref,
        remark2=original_package.remark2,
        remark_of_hotels=original_package.remark_of_hotels,
        special_note=original_package.special_note,
        tour_pack_type=original_package.tour_pack_type,
        commission_rate_hotel=original_package.commission_rate_hotel,
        commission_rate_services=original_package.commission_rate_services,
        hotel_costs=original_package.hotel_costs,
        discounts=original_package.discounts,
        extra_costs=original_package.extra_costs,
        prepare_by_user=request.user
    )

    # Copy tour days and their associated services
    for original_day in original_package.tour_days.all():
        new_day = TourDay.objects.create(
            tour_package=new_package,
            date=original_day.date,
            city=original_day.city,
            hotel=original_day.hotel
        )

        # Copy services
        for original_service in original_day.services.all():
            TourDayService.objects.create(
                tour_day=new_day,
                service=original_service.service,
                price_at_booking=original_service.price_at_booking
            )

        # Copy guide services
        for original_guide_service in original_day.guide_services.all():
            TourDayGuideService.objects.create(
                tour_day=new_day,
                guide_service=original_guide_service.guide_service,
                price_at_booking=original_guide_service.price_at_booking
            )

    # Recalculate totals for the new package
    service_grand_total, hotel_grand_total, grand_total, total_discount, total_extra_cost = calculate_totals(new_package)

    new_package.service_grand_total = service_grand_total
    new_package.hotel_grand_total = hotel_grand_total
    new_package.grand_total_cost = grand_total

    # Recalculate commission amounts
    total_room_nights = sum(
        safe_decimal(hotel.get('room')) * safe_decimal(hotel.get('nights'))
        for hotel in new_package.hotel_costs
    )

    new_package.commission_amount_hotel = new_package.commission_rate_hotel * total_room_nights
    new_package.commission_amount_services = new_package.commission_rate_services * new_package.service_grand_total / Decimal('100')

    new_package.save()

    messages.success(request, f"Tour package '{new_package.name}' has been created as a copy.")
    return redirect('tour_package_edit', package_reference=new_package.package_reference)


@login_required
def service_price_form(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            with transaction.atomic():
                # Get or create service type
                service_type, _ = ServiceType.objects.get_or_create(
                    name=data['service_type']
                )

                # Create service
                service = Service.objects.create(
                    name=data['name'],
                    service_type=service_type
                )

                # Add cities
                for city_id in data['cities']:
                    city = City.objects.get(id=city_id)
                    service.cities.add(city)

                # Create prices
                for price_data in data['prices']:
                    ServicePrice.objects.create(
                        service=service,
                        tour_pack_type_id=price_data['tour_pack_type'],
                        price=price_data['price']
                    )

            return JsonResponse({'status': 'success', 'message': 'Service created successfully'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # GET request - render form
    cities = City.objects.all()
    tour_pack_types = TourPackType.objects.all()
    service_types = ServiceType.objects.all()

    context = {
        'cities': cities,
        'tour_pack_types': tour_pack_types,
        'service_types': service_types,
    }
    return render(request, 'tour_quote/service_price_form.html', context)

@login_required
def service_price_edit(request):
    """
    View for editing service prices.
    """
    services = Service.objects.all().select_related('service_type')
    tour_pack_types = TourPackType.objects.all()

    # Convert services to JSON for safe handling of special characters
    services_list = [{
        'id': service.id,
        'name': service.name,
        'service_type': service.service_type.name if service.service_type else ''
    } for service in services]
    services_json = json.dumps(services_list)

    context = {
        'services': services,
        'tour_pack_types': tour_pack_types,
        'services_json': services_json,
    }
    return render(request, 'tour_quote/service_price_edit.html', context)

@login_required
@require_http_methods(['GET'])
def get_service_prices(request, service_id):
    """
    API endpoint to get prices for a specific service.
    Returns empty string for blank prices, but actual 0 for zero prices.
    """
    try:
        service = Service.objects.get(id=service_id)
        prices = ServicePrice.objects.filter(service=service).select_related('tour_pack_type')

        return JsonResponse({
            'service': {
                'id': service.id,
                'name': service.name,
                'service_type': service.service_type.name,
            },
            'prices': [{
                'id': price.id,
                'tour_pack_type_id': price.tour_pack_type.id,
                'tour_pack_type_name': price.tour_pack_type.name,
                'price': str(price.price) if price.price is not None else ''  # Keep 0 as "0", blank as ""
            } for price in prices]
        })
    except Service.DoesNotExist:
        return JsonResponse({'error': 'Service not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_http_methods(['POST'])
def save_service_prices(request):
    """
    API endpoint to save service prices.
    - Accepts 0 as a valid price
    - Skips creation/removes existing prices for blank inputs
    """
    try:
        data = json.loads(request.body)
        service_id = data['service_id']
        prices = data['prices']
        service = Service.objects.get(id=service_id)

        with transaction.atomic():
            processed_price_ids = []

            for price_data in prices:
                tour_pack_type_id = price_data['tour_pack_type_id']
                price_id = price_data.get('price_id')
                price_input = price_data['price']

                # Handle blank input case
                if price_input == '' or price_input is None:
                    # If there's an existing price record, delete it
                    if price_id:
                        ServicePrice.objects.filter(id=price_id).delete()
                    continue

                # Convert price to Decimal
                try:
                    price_value = Decimal(str(price_input))
                except (ValueError, TypeError, InvalidOperation):
                    continue

                # Handle existing price record
                if price_id:
                    try:
                        price_obj = ServicePrice.objects.get(id=price_id)
                        price_obj.price = price_value
                        price_obj.save()
                        processed_price_ids.append(price_obj.id)
                    except ServicePrice.DoesNotExist:
                        # If price record doesn't exist anymore, create new if not blank
                        price_obj = ServicePrice.objects.create(
                            service=service,
                            tour_pack_type_id=tour_pack_type_id,
                            price=price_value
                        )
                        processed_price_ids.append(price_obj.id)
                else:
                    # Create new price record
                    price_obj = ServicePrice.objects.create(
                        service=service,
                        tour_pack_type_id=tour_pack_type_id,
                        price=price_value
                    )
                    processed_price_ids.append(price_obj.id)

            # Delete any price records that weren't processed
            ServicePrice.objects.filter(
                service=service
            ).exclude(
                id__in=processed_price_ids
            ).delete()

            # Get updated prices for response
            updated_prices = ServicePrice.objects.filter(
                service=service
            ).select_related('tour_pack_type')

            return JsonResponse({
                'status': 'success',
                'message': 'Prices updated successfully',
                'prices': [{
                    'id': price.id,
                    'tour_pack_type_id': price.tour_pack_type.id,
                    'tour_pack_type_name': price.tour_pack_type.name,
                    'price': str(price.price) if price.price is not None else ''
                } for price in updated_prices]
            })

    except Service.DoesNotExist:
        return JsonResponse({'error': 'Service not found'}, status=404)
    except Exception as e:
        print(f"Error in save_service_prices: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def service_list(request):
    """
    View to list all services with their prices.
    """
    services = Service.objects.all().select_related('service_type').prefetch_related('prices__tour_pack_type')
    context = {
        'services': services
    }
    return render(request, 'tour_quote/service_list.html', context)

@login_required
def export_tour_package_json(request, pk):
    """
    Exports a single TourPackageQuote and all its related data into a JSON file.
    This view manually constructs a dictionary to ensure all related data (days, hotels, costs)
    is included in a clean, nested format.
    """
    # Retrieve the specific tour package or return a 404 error if not found
    tour_package = get_object_or_404(TourPackageQuote, pk=pk)

    # Manually build the dictionary with all the required data
    package_data = {
        "id": tour_package.id,
        "name": tour_package.name,
        "customer_name": tour_package.customer_name,
        "created_at": tour_package.created_at.isoformat(),
        "package_reference": tour_package.package_reference,
        "grand_total_cost": str(tour_package.grand_total_cost),
        "service_grand_total": str(tour_package.service_grand_total),
        "hotel_grand_total": str(tour_package.hotel_grand_total),
        "remark": tour_package.remark,
        "connection_ref": tour_package.connection_ref,
        "remark2": tour_package.remark2,
        "remark_of_hotels": tour_package.remark_of_hotels,
        "special_note": tour_package.special_note,
        "tour_pack_type": tour_package.tour_pack_type.name if tour_package.tour_pack_type else None,
        "hotel_costs": tour_package.hotel_costs,
        "discounts": tour_package.discounts,
        "extra_costs": tour_package.extra_costs,
        "commission_info": {
            "hotel_rate": str(tour_package.commission_rate_hotel),
            "hotel_amount": str(tour_package.commission_amount_hotel),
            "services_rate": str(tour_package.commission_rate_services),
            "services_amount": str(tour_package.commission_amount_services),
            "total": str(tour_package.commission_amount_hotel + tour_package.commission_amount_services)
        },
        "tour_days": []
    }

    # Add related tour days to the dictionary
    for day in tour_package.tour_days.all():
        day_data = {
            "date": day.date.isoformat(),
            "city": day.city.name,
            "hotel": day.hotel.name,
            "services": [],
            "guide_services": []
        }

        # Add services for this day
        for service in day.services.all():
            day_data["services"].append({
                "name": service.service.name,
                "service_type": service.service.service_type.name,
                "price_at_booking": str(service.price_at_booking)
            })

        # Add guide services for this day
        for guide_service in day.guide_services.all():
            day_data["guide_services"].append({
                "name": guide_service.guide_service.name,
                "price_at_booking": str(guide_service.price_at_booking)
            })

        package_data["tour_days"].append(day_data)

    # Create a JSON response. The `Content-Disposition` header tells the
    # browser to treat the response as a file download.
    response = JsonResponse(package_data, json_dumps_params={'indent': 2})
    response['Content-Disposition'] = f'attachment; filename="tour_package_{tour_package.pk}.json"'

    return response

@login_required
def import_tour_package_json(request):
    """
    Import a tour package from a JSON file and create a new TourPackageQuote.
    """
    if request.method == 'POST' and request.FILES.get('json_file'):
        try:
            # Read the uploaded JSON file
            json_file = request.FILES['json_file']
            json_data = json.loads(json_file.read().decode('utf-8'))

            # Extract basic package information
            with transaction.atomic():
                # Create the tour package
                tour_package = TourPackageQuote(
                    name=json_data.get('name', 'Imported Package'),
                    customer_name=json_data.get('customer_name', 'Imported Customer'),
                    remark=json_data.get('remark'),
                    connection_ref=json_data.get('connection_ref'),
                    remark2=json_data.get('remark2'),
                    remark_of_hotels=json_data.get('remark_of_hotels'),
                    special_note=json_data.get('special_note'),
                    hotel_costs=json_data.get('hotel_costs', []),
                    discounts=json_data.get('discounts', []),
                    extra_costs=json_data.get('extra_costs', []),
                    grand_total_cost=Decimal(json_data.get('grand_total_cost', '0')),
                    service_grand_total=Decimal(json_data.get('service_grand_total', '0')),
                    hotel_grand_total=Decimal(json_data.get('hotel_grand_total', '0')),
                    commission_rate_hotel=Decimal(json_data.get('commission_info', {}).get('hotel_rate', '0')),
                    commission_amount_hotel=Decimal(json_data.get('commission_info', {}).get('hotel_amount', '0')),
                    commission_rate_services=Decimal(json_data.get('commission_info', {}).get('services_rate', '0')),
                    commission_amount_services=Decimal(json_data.get('commission_info', {}).get('services_amount', '0')),
                    prepare_by_user=request.user
                )

                # Set tour pack type if it exists
                if json_data.get('tour_pack_type'):
                    tour_pack_type, _ = TourPackType.objects.get_or_create(
                        name=json_data.get('tour_pack_type')
                    )
                    tour_package.tour_pack_type = tour_pack_type

                # Save the package to generate a reference ID
                tour_package.save()

                # Process tour days if they exist
                if json_data.get('tour_days'):
                    for day_data in json_data['tour_days']:
                        # Get or create city
                        city, _ = City.objects.get_or_create(name=day_data.get('city'))

                        # Get or create hotel
                        hotel, _ = Hotel.objects.get_or_create(
                            name=day_data.get('hotel'),
                            city=city
                        )

                        # Create tour day
                        tour_day = TourDay.objects.create(
                            tour_package=tour_package,
                            date=datetime.fromisoformat(day_data.get('date')).date(),
                            city=city,
                            hotel=hotel
                        )

                        # Add services
                        if day_data.get('services'):
                            for service_data in day_data['services']:
                                # Get or create service type
                                service_type, _ = ServiceType.objects.get_or_create(
                                    name=service_data.get('service_type', 'General')
                                )

                                # Get or create service
                                service, _ = Service.objects.get_or_create(
                                    name=service_data.get('name'),
                                    service_type=service_type
                                )
                                service.cities.add(city)

                                # Create tour day service
                                TourDayService.objects.create(
                                    tour_day=tour_day,
                                    service=service,
                                    price_at_booking=Decimal(service_data.get('price_at_booking', '0'))
                                )

                        # Add guide services
                        if day_data.get('guide_services'):
                            for guide_service_data in day_data['guide_services']:
                                # Get or create guide service
                                guide_service, _ = GuideService.objects.get_or_create(
                                    name=guide_service_data.get('name')
                                )

                                # Create tour day guide service
                                TourDayGuideService.objects.create(
                                    tour_day=tour_day,
                                    guide_service=guide_service,
                                    price_at_booking=Decimal(guide_service_data.get('price_at_booking', '0'))
                                )

            messages.success(request, f"Successfully imported tour package: {tour_package.name}")
            return redirect('tour_package_detail', package_reference=tour_package.package_reference)

        except json.JSONDecodeError:
            messages.error(request, "Invalid JSON file. Please upload a valid JSON file.")
        except Exception as e:
            messages.error(request, f"Error importing tour package: {str(e)}")

    return render(request, 'tour_quote/import_tour_package.html')

@login_required
@require_http_methods(["GET"])
def export_tourday_excel(request, pk):
    """
    Export tour days and services to Excel file.
    Format:
    - Row 1: Headers
    - Row 2: First blank row
    - Row 3: Second blank row with black background
    - Row 4: Tour pack type + tour quote name
    - Subsequent rows: Services grouped by hotel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from collections import defaultdict, OrderedDict
    from datetime import timedelta
    import re

    package = get_object_or_404(TourPackageQuote, pk=pk)

    # Define Office 2013-2022 Theme XML
    OFFICE_THEME_XML = """
<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="Office Theme">
   <a:themeElements>
      <a:clrScheme name="Office">
         <a:dk1><a:sysClr val="windowText" lastClr="000000"/></a:dk1>
         <a:lt1><a:sysClr val="window" lastClr="FFFFFF"/></a:lt1>
         <a:dk2><a:srgbClr val="44546A"/></a:dk2>
         <a:lt2><a:srgbClr val="E7E6E6"/></a:lt2>
         <a:accent1><a:srgbClr val="4472C4"/></a:accent1>
         <a:accent2><a:srgbClr val="ED7D31"/></a:accent2>
         <a:accent3><a:srgbClr val="A5A5A5"/></a:accent3>
         <a:accent4><a:srgbClr val="FFC000"/></a:accent4>
         <a:accent5><a:srgbClr val="5B9BD5"/></a:accent5>
         <a:accent6><a:srgbClr val="70AD47"/></a:accent6>
         <a:hlink><a:srgbClr val="0563C1"/></a:hlink>
         <a:folHlink><a:srgbClr val="954F72"/></a:folHlink>
      </a:clrScheme>
      <a:fontScheme name="Office">
         <a:majorFont>
            <a:latin typeface="Calibri Light" panose="020F0302020204030204"/>
            <a:ea typeface=""/>
            <a:cs typeface=""/>
         </a:majorFont>
         <a:minorFont>
            <a:latin typeface="Calibri" panose="020F0502020204030204"/>
            <a:ea typeface=""/>
            <a:cs typeface=""/>
         </a:minorFont>
      </a:fontScheme>
      <a:fmtScheme name="Office">
         <a:fillStyleLst>
            <a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
            <a:gradFill rotWithShape="1"><a:gsLst><a:gs pos="0"><a:schemeClr val="phClr"><a:tint val="50000"/><a:satMod val="300000"/></a:schemeClr></a:gs><a:gs pos="35000"><a:schemeClr val="phClr"><a:tint val="37000"/><a:satMod val="300000"/></a:schemeClr></a:gs><a:gs pos="100000"><a:schemeClr val="phClr"><a:tint val="15000"/><a:satMod val="350000"/></a:schemeClr></a:gs></a:gsLst><a:lin ang="16200000" scaled="1"/></a:gradFill>
            <a:gradFill rotWithShape="1"><a:gsLst><a:gs pos="0"><a:schemeClr val="phClr"><a:shade val="51000"/><a:satMod val="130000"/></a:schemeClr></a:gs><a:gs pos="80000"><a:schemeClr val="phClr"><a:shade val="93000"/><a:satMod val="130000"/></a:schemeClr></a:gs><a:gs pos="100000"><a:schemeClr val="phClr"><a:shade val="94000"/><a:satMod val="135000"/></a:schemeClr></a:gs></a:gsLst><a:lin ang="16200000" scaled="0"/></a:gradFill>
         </a:fillStyleLst>
         <a:lnStyleLst>
            <a:ln w="9525" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"><a:shade val="95000"/><a:satMod val="105000"/></a:schemeClr></a:solidFill><a:prstDash val="solid"/></a:ln>
            <a:ln w="25400" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/></a:ln>
            <a:ln w="38100" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/></a:ln>
         </a:lnStyleLst>
         <a:effectStyleLst>
            <a:effectStyle><a:effectLst><a:outerShdw blurRad="40000" dist="20000" dir="5400000" rotWithShape="0"><a:srgbClr val="000000"><a:alpha val="38000"/></a:srgbClr></a:outerShdw></a:effectLst></a:effectStyle>
            <a:effectStyle><a:effectLst><a:outerShdw blurRad="40000" dist="23000" dir="5400000" rotWithShape="0"><a:srgbClr val="000000"><a:alpha val="35000"/></a:srgbClr></a:outerShdw></a:effectLst></a:effectStyle>
            <a:effectStyle><a:effectLst><a:outerShdw blurRad="40000" dist="23000" dir="5400000" rotWithShape="0"><a:srgbClr val="000000"><a:alpha val="35000"/></a:srgbClr></a:outerShdw></a:effectLst><a:scene3d><a:camera prst="orthographicFront"><a:rot lat="0" lon="0" rev="0"/></a:camera><a:lightRig rig="threePt" dir="t"><a:rot lat="0" lon="0" rev="1200000"/></a:lightRig></a:scene3d><a:sp3d><a:bevelT w="63500" h="25400"/></a:sp3d></a:effectStyle>
         </a:effectStyleLst>
         <a:bgFillStyleLst>
            <a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
            <a:gradFill rotWithShape="1"><a:gsLst><a:gs pos="0"><a:schemeClr val="phClr"><a:tint val="40000"/><a:satMod val="350000"/></a:schemeClr></a:gs><a:gs pos="40000"><a:schemeClr val="phClr"><a:tint val="45000"/><a:shade val="99000"/><a:satMod val="350000"/></a:schemeClr></a:gs><a:gs pos="100000"><a:schemeClr val="phClr"><a:shade val="20000"/><a:satMod val="255000"/></a:schemeClr></a:gs></a:gsLst><a:path path="circle"><a:fillToRect l="50000" t="-80000" r="50000" b="180000"/></a:path></a:gradFill>
            <a:gradFill rotWithShape="1"><a:gsLst><a:gs pos="0"><a:schemeClr val="phClr"><a:tint val="80000"/><a:satMod val="300000"/></a:schemeClr></a:gs><a:gs pos="100000"><a:schemeClr val="phClr"><a:shade val="30000"/><a:satMod val="200000"/></a:schemeClr></a:gs></a:gsLst><a:path path="circle"><a:fillToRect l="50000" t="50000" r="50000" b="50000"/></a:path></a:gradFill>
         </a:bgFillStyleLst>
      </a:fmtScheme>
   </a:themeElements>
   <a:objectDefaults/>
   <a:extraClrSchemeLst/>
</a:theme>"""

    # Create workbook
    wb = Workbook()
    # Apply the Office 2013-2022 theme
    wb.loaded_theme = OFFICE_THEME_XML
    ws = wb.active
    ws.title = "Tour Days Export"

    # Row 1: Headers
    headers = [
        'Ref nr.', 'Pax', 'Arr.', 'Dep.', 'Nt', 'Detail', 'P.U.', 'P.U.Time', 'D.O.',
        'Flight / Train / Boat / others', '', 'INVOICE NR & TOTAL', 'INVOICE to Connections',
        'Promotion', 'booking status \\ hotel cfrm nr', 'INVOICE \\ EXPENSES by supplier',
        '', 'due date', 'supplier \\ guide', 'PROFIT PER PRODUCT', 'PROFIT on file'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

        # Center align headers from 'Ref nr.' to 'D.O.' (first 9 columns) EXCEPT 'Detail' (col 6)
        if col <= 9 and col != 6:
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        else:
            cell.alignment = Alignment(wrap_text=True)

    # Row 2: First blank row (empty)
    # Row 3: Second blank row with separator background (Black)
    separator_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    for col in range(1, 22):
        ws.cell(row=3, column=col).fill = separator_fill

    # Freeze panes - freeze row 1 (header) so rows below scroll
    ws.freeze_panes = 'A3'

    # Track current row for dynamic placement
    current_info_row = 4

    # Populate Mo no. at J4
    mo_no_value = f"Mo no.:{package.package_reference}" if package.package_reference else "Mo no.:"
    cell = ws.cell(row=4, column=10, value=mo_no_value)
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=4, column=12, value="Invoice nr")
    ws.cell(row=4, column=13, value="Bkg nr")

    # Add Remark 1 row only if remark exists
    # if package.remark:
    #     ws.cell(row=current_info_row, column=3, value=package.remark)
    #     current_info_row += 1

    # Connection Ref + Tour pack type + tour quote name (all in same row)
    connection_ref_value = package.connection_ref if package.connection_ref else ''
    tour_pack_type_value = package.tour_pack_type.name if package.tour_pack_type else ''

    # Extract number only for Pax column
    pax_value = tour_pack_type_value
    if tour_pack_type_value:
        match = re.search(r'(\d+)', tour_pack_type_value)
        if match:
            pax_value = int(match.group(1))

    ws.cell(row=current_info_row, column=1, value=connection_ref_value)  # Ref nr.
    cell = ws.cell(row=current_info_row, column=2, value=pax_value)  # Tour Package Type (Pax)
    cell.alignment = Alignment(horizontal='center')
    billing_name_value = package.billing_name.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ') if package.billing_name else ''
    ws.cell(row=current_info_row, column=3, value=billing_name_value or package.name)  # Tour quote name
    cell = ws.cell(row=current_info_row, column=7, value=package.customer_name)  # Customer Name at G5 (relative)
    cell.alignment = Alignment(horizontal='center')
    current_info_row += 1

    # Add Special Note rows if exists
    if package.special_note:
        special_notes = package.special_note.split('\n')
        for note in special_notes:
            if note.strip():  # Only add non-empty lines
                ws.cell(row=current_info_row, column=3, value=note.strip())
                current_info_row += 1

    # Determine data start row based on pax count
    # If tour pack type > 2pax, add one more blank row
    data_start_row = current_info_row
    if package.tour_pack_type:
       data_start_row = current_info_row + 1  # Add one more blank row

    # Service type priority order
    SERVICE_TYPE_ORDER = {
        'transfer': 1,
        'package': 3,
        'tour': 4,
        'custom': 5,
        'zero': 6,
    }

    def get_service_type_order(service_type_name, service_name=''):
        """Get sort order for service type.
        Special case: any service with ** in name is treated as transfer (order=1) - appears before hotel
        """
        # Special case: any service with ** in name acts as transfer (before hotel)
        if '**' in service_name:
            return 1  # Same as transfer

        name_lower = service_type_name.lower() if service_type_name else ''
        for key, order in SERVICE_TYPE_ORDER.items():
            if key in name_lower:
                return order
        return 99  # Unknown types go last

    # Get all tour days ordered by date
    tour_days = list(package.tour_days.all().order_by('date'))

    # Build hotel info lookup from hotel_costs (name -> list of room configs)
    # hotel_costs contains: name, type, room, extraBedPrice, price, nights, promotion
    # Same hotel can have multiple room types with different prices
    hotel_info_lookup = {}
    for cost in (package.hotel_costs or []):
        hotel_name = cost.get('name', '')
        room_count = cost.get('room', 1)
        room_type = cost.get('type', '')
        extra_bed_price = cost.get('extraBedPrice')
        room_price = cost.get('price', 0)
        nights = cost.get('nights', 1)
        promotion = cost.get('promotion', '')

        # Build display string: "3x Premier room + extra bed"
        room_part = ""
        if room_count and room_type:
            room_part = f"{room_count}x {room_type}"
        elif room_type:
            room_part = room_type

        # Build price formula: (room_price + extra_bed_price) * nights * rooms
        extra_bed = float(extra_bed_price) if extra_bed_price and float(extra_bed_price) > 0 else 0
        room_price_val = float(room_price) if room_price else 0
        nights_val = int(nights) if nights else 1
        rooms_val = int(room_count) if room_count else 1

        # Formula string: ((room_price * rooms) + extra_bed) * nights
        if extra_bed > 0:
            price_formula = f"=(({room_price_val}*{rooms_val})+{extra_bed})*{nights_val}"
            display_name = f"{room_part} + extra bed" if room_part else ""
        else:
            price_formula = f"={room_price_val}*{rooms_val}*{nights_val}"
            display_name = room_part

        # Store as list to handle same hotel with multiple room types
        if hotel_name not in hotel_info_lookup:
            hotel_info_lookup[hotel_name] = []
        hotel_info_lookup[hotel_name].append({
            'display': display_name,
            'price_formula': price_formula,
            'promotion': promotion,
            'date_str': cost.get('date', ''),
            'nights': nights_val
        })

    # Group tour days by hotel to calculate hotel stays
    # Use list to handle non-contiguous stays at same hotel
    hotel_groups_list = []
    last_hotel_name = None
    current_group = None

    for day in tour_days:
        hotel_name = day.hotel.name if day.hotel else 'No Hotel'

        if hotel_name != last_hotel_name:
            current_group = {
                'hotel_name': hotel_name,
                'arrival_date': day.date,
                'departure_date': day.date + timedelta(days=1),  # Checkout is next day
                'nights': 1,
                'services': [],
                'guide_services': [],
                'dates': [day.date],
            }
            hotel_groups_list.append(current_group)
            last_hotel_name = hotel_name
        else:
            # Extend the stay - update departure date and nights
            current_group['departure_date'] = day.date + timedelta(days=1)
            current_group['nights'] += 1
            current_group['dates'].append(day.date)

        # Collect services for this day under this hotel
        for service in day.services.all():
            service_type_name = service.service.service_type.name if service.service.service_type else ''
            service_name = service.service.name
            current_group['services'].append({
                'date': day.date,
                'service_type': service_type_name,
                'sort_order': get_service_type_order(service_type_name, service_name),
                'name': service_name,
                'price': service.price_at_booking,
            })

        # Collect guide services
        for guide_service in day.guide_services.all():
            current_group['guide_services'].append({
                'date': day.date,
                'service_type': 'custom',
                'sort_order': 5,
                'name': guide_service.guide_service.name,
                'price': guide_service.price_at_booking,
            })

    # Build export items grouped by hotel
    final_items = []

    for hotel_data in hotel_groups_list:
        hotel_name = hotel_data['hotel_name']
        # Combine regular services and guide services, then sort by date
        regular_services = hotel_data['services']
        guide_services = hotel_data['guide_services']

        # Combine all services and sort by date to maintain date sequence
        all_services = regular_services + guide_services
        all_services.sort(key=lambda x: x['date'])

        # Hotel rows - create one row per room configuration (same hotel can have multiple room types)
        hotel_rows = []
        if 'No Hotel' in hotel_name:
            pass
        elif hotel_name in hotel_info_lookup:
            room_configs = hotel_info_lookup[hotel_name]

            # Filter configs by date matching - check if config date matches ANY day in the stay
            matching_configs = []
            group_dates = hotel_data.get('dates', [hotel_data['arrival_date']])

            for config in room_configs:
                c_date = config.get('date_str', '')
                if not c_date:
                    continue

                # Check against all dates in the group
                matched_date = None
                for d in group_dates:
                    day_str_pad = f"{d.day:02d}"
                    month_str = d.strftime('%b').lower()

                    # Check if starts with day (padded or not) and contains month (case-insensitive)
                    if (c_date.startswith(day_str_pad) or c_date.startswith(str(d.day))) and month_str in c_date.lower():
                        matched_date = d
                        break

                if matched_date:
                    # Create a copy of config with the specific arrival date
                    config_copy = config.copy()
                    config_copy['specific_arrival_date'] = matched_date
                    matching_configs.append(config_copy)

            # Fallback if no specific date match found
            if not matching_configs:
                 matching_configs.append({
                     'display': '',
                     'price_formula': None,
                     'promotion': ''
                 })

            # Sort configs by date to ensure they appear in order, even if input was out of order
            matching_configs.sort(key=lambda x: x.get('specific_arrival_date') or hotel_data['arrival_date'])

            for room_config in matching_configs:
                hotel_display_name = hotel_name
                if room_config.get('display'):
                    hotel_display_name = f"{hotel_name}, {room_config['display']}"

                # Determine dates and nights
                arrival_date = room_config.get('specific_arrival_date', hotel_data['arrival_date'])
                nights = room_config.get('nights', hotel_data['nights'])

                # Calculate departure date based on specific arrival and nights
                try:
                     # ensure nights is valid number
                    nights_int = int(nights) if nights else 0
                    departure_date = arrival_date + timedelta(days=nights_int)
                except:
                    departure_date = hotel_data['departure_date']

                hotel_rows.append({
                    'arrival_date': arrival_date,
                    'departure_date': departure_date,
                    'nights': nights,
                    'service_name': hotel_display_name,
                    'price': None,
                    'price_formula': room_config.get('price_formula'),
                    'promotion': room_config.get('promotion', ''),
                    'is_hotel': True,
                    'supplier_guide': hotel_name,
                })
        else:
            # No room config found, just use hotel name
            hotel_rows.append({
                'arrival_date': hotel_data['arrival_date'],
                'departure_date': hotel_data['departure_date'],
                'nights': hotel_data['nights'],
                'service_name': hotel_name,
                'price': None,
                'price_formula': None,
                'is_hotel': True,
                'supplier_guide': hotel_name,
            })

        # Check if service should come before hotel (transfer type or ** in name)
        def should_be_before_hotel(svc):
            service_type = svc.get('service_type', '').lower()
            service_name = svc.get('name', '')
            return 'transfer' in service_type or '**' in service_name or 'transfer' in service_name.lower()

        # Find the index of the last service that should be before hotel
        last_before_hotel_idx = -1
        for idx, svc in enumerate(all_services):
            if should_be_before_hotel(svc):
                last_before_hotel_idx = idx

        if last_before_hotel_idx >= 0:
            # Case 1: Has transfers or ** services
            for svc in all_services[:last_before_hotel_idx + 1]:
                final_items.append({
                    'arrival_date': svc['date'],
                    'departure_date': None,
                    'nights': None,
                    'service_name': svc['name'],
                    'price': svc['price'],
                })
            # Add all hotel rows (multiple room types)
            for hotel_row in hotel_rows:
                final_items.append(hotel_row)
            for svc in all_services[last_before_hotel_idx + 1:]:
                final_items.append({
                    'arrival_date': svc['date'],
                    'departure_date': None,
                    'nights': None,
                    'service_name': svc['name'],
                    'price': svc['price'],
                })
        else:
            # Case 2: No transfers or ** services
            for hotel_row in hotel_rows:
                final_items.append(hotel_row)
            for svc in all_services:
                final_items.append({
                    'arrival_date': svc['date'],
                    'departure_date': None,
                    'nights': None,
                    'service_name': svc['name'],
                    'price': svc['price'],
                })

    # Fill column 11 (separator) with separator fill from row 1 up to data_start_row
    for r in range(1, data_start_row):
        ws.cell(row=r, column=11).fill = separator_fill

    # Write data rows from data_start_row
    current_row = data_start_row

    # Add extra_costs at the top (before services)
    extra_costs = package.extra_costs or []
    for extra_cost in extra_costs:
        cost_name = extra_cost.get('item', '') or extra_cost.get('name', '')
        cost_amount = extra_cost.get('amount', 0)

        if cost_name:
            ws.cell(row=current_row, column=1, value='')
            ws.cell(row=current_row, column=2, value='')
            ws.cell(row=current_row, column=3, value='')
            ws.cell(row=current_row, column=4, value='')
            ws.cell(row=current_row, column=5, value='')
            ws.cell(row=current_row, column=6, value=cost_name)
            ws.cell(row=current_row, column=11).fill = separator_fill

            if cost_amount:
                cell = ws.cell(row=current_row, column=13, value=float(cost_amount))
                cell.number_format = '#,##0.00'
            else:
                 ws.cell(row=current_row, column=13, value='')

            # Column T (20) Profit Formula: M - P - Q
            ws.cell(row=current_row, column=20, value=f"=M{current_row}-P{current_row}-Q{current_row}")
            ws.cell(row=current_row, column=20).number_format = '#,##0.00'

            # Format P, Q, R
            ws.cell(row=current_row, column=16).number_format = '#,##0.00'
            ws.cell(row=current_row, column=17).number_format = '#,##0.00'
            ws.cell(row=current_row, column=18).number_format = 'dd-mmm-yy'

            current_row += 1

    # Add discounts after extra_costs
    discounts = package.discounts or []
    red_font = Font(color="FF0000") # Red color

    for discount in discounts:
        discount_name = discount.get('item', '') or discount.get('name', '')
        discount_amount = discount.get('amount', 0)

        if discount_name:
            ws.cell(row=current_row, column=1, value='')
            ws.cell(row=current_row, column=2, value='')
            ws.cell(row=current_row, column=3, value='')
            ws.cell(row=current_row, column=4, value='')
            ws.cell(row=current_row, column=5, value='')

            # Detail column with red font
            cell = ws.cell(row=current_row, column=6, value=discount_name)
            cell.font = red_font

            ws.cell(row=current_row, column=11).fill = separator_fill

            if discount_amount:
                # Ensure negative amount
                amount_val = -abs(float(discount_amount))
                cell = ws.cell(row=current_row, column=13, value=amount_val)
                cell.number_format = '#,##0.00'
                cell.font = red_font
            else:
                 ws.cell(row=current_row, column=13, value='')

            # Column T (20) Profit Formula: M - P - Q
            ws.cell(row=current_row, column=20, value=f"=M{current_row}-P{current_row}-Q{current_row}")
            ws.cell(row=current_row, column=20).number_format = '#,##0.00'

            # Format P, Q, R
            ws.cell(row=current_row, column=16).number_format = '#,##0.00'
            ws.cell(row=current_row, column=17).number_format = '#,##0.00'
            ws.cell(row=current_row, column=18).number_format = 'dd-mmm-yy'

            current_row += 1

    # Define alignment styles
    center_align = Alignment(horizontal='center', vertical='center')

    for item in final_items:
        # Ref nr. (A) - Col 1
        cell = ws.cell(row=current_row, column=1, value='')
        cell.alignment = center_align

        # Pax (B) - Col 2
        cell = ws.cell(row=current_row, column=2, value='')
        cell.alignment = center_align

        if item['arrival_date']:
            cell = ws.cell(row=current_row, column=3, value=item['arrival_date'])
            cell.number_format = 'dd-mmm-yy'
            cell.alignment = center_align
        else:
            cell = ws.cell(row=current_row, column=3, value='')
            cell.alignment = center_align

        if item['departure_date']:
            cell = ws.cell(row=current_row, column=4, value=item['departure_date'])
            cell.number_format = 'dd-mmm-yy'
            cell.alignment = center_align
        else:
            cell = ws.cell(row=current_row, column=4, value='')
            cell.alignment = center_align

        nights_str = item['nights'] if item['nights'] else ''
        ws.cell(row=current_row, column=5, value=nights_str)
        ws.cell(row=current_row, column=6, value=item['service_name'])
        # Columns 7-10 are empty (P.U., P.U.Time, D.O., Flight/Train/Boat/others)

        # Apply alignment to columns G(7), H(8), I(9), J(10) which are currently empty in data but might need alignment if filled?
        # Wait, the code below says "Columns 7-10 are empty" in the comment, but looking at previous lines,
        # I don't see them being filled from 'item'.
        # Let's check the 'item' structure. It has 'service_name', 'price', dates.
        # It seems the export DOES NOT populate P.U., Time, D.O., Flight info from the item dictionary yet.
        # But the user asked to align them. They might be manually filling them later or I missed where they are filled.
        # The code comments say: "# Columns 7-10 are empty (P.U., P.U.Time, D.O., Flight/Train/Boat/others)"
        # So I will just apply the alignment to the empty cells so that if they type in it, it is aligned.

        for col in [7, 8, 9, 10]:
            cell = ws.cell(row=current_row, column=col)
            cell.alignment = center_align

        # Column 11 is separator (separator fill)
        ws.cell(row=current_row, column=11).fill = separator_fill

        # Column 12 is INVOICE NR & TOTAL (empty)

        if item.get('is_hotel') and item.get('price_formula'):
            cell = ws.cell(row=current_row, column=13, value=item['price_formula'])
            cell.number_format = '#,##0.00'
            if item.get('promotion'):
                ws.cell(row=current_row, column=14, value=item['promotion'])
        else:
            if item['price']:
                cell = ws.cell(row=current_row, column=13, value=float(item['price']))
                cell.number_format = '#,##0.00'
            else:
                ws.cell(row=current_row, column=13, value='')

        if item.get('supplier_guide'):
            ws.cell(row=current_row, column=19, value=item['supplier_guide'])

        # Column T (20) Profit Formula: M - P - Q
        ws.cell(row=current_row, column=20, value=f"=M{current_row}-P{current_row}-Q{current_row}")
        ws.cell(row=current_row, column=20).number_format = '#,##0.00'

        # Format P, Q, R
        ws.cell(row=current_row, column=16).number_format = '#,##0.00'
        ws.cell(row=current_row, column=17).number_format = '#,##0.00'
        ws.cell(row=current_row, column=18).number_format = 'dd-mmm-yy'

        current_row += 1

    # Add 3 blank rows separator after services
    # Blank row 1
    ws.cell(row=current_row, column=11).fill = separator_fill
    current_row += 1

    # Blank row 2
    ws.cell(row=current_row, column=11).fill = separator_fill
    current_row += 1

    # Blank row 3 with separator fill
    for col in range(1, 22):
        ws.cell(row=current_row, column=col).fill = separator_fill
    current_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # Ref nr.
    ws.column_dimensions['B'].width = 8   # Pax
    ws.column_dimensions['C'].width = 12  # Arr.
    ws.column_dimensions['D'].width = 12  # Dep.
    ws.column_dimensions['E'].width = 5   # Nt
    ws.column_dimensions['F'].width = 50  # Detail
    ws.column_dimensions['G'].width = 25   # P.U.
    ws.column_dimensions['H'].width = 10  # P.U.Time
    ws.column_dimensions['I'].width = 25   # D.O.
    ws.column_dimensions['J'].width = 25  # Flight / Train / Boat / others
    ws.column_dimensions['K'].width = 1   # Separator (black border)
    ws.column_dimensions['L'].width = 12  # INVOICE NR & TOTAL
    ws.column_dimensions['M'].width = 18  # INVOICE to Connections
    ws.column_dimensions['N'].width = 20  # remarks
    ws.column_dimensions['O'].width = 20  # booking status / hotel cfrm nr
    ws.column_dimensions['P'].width = 20  # INVOICE by supplier / EXPENSES to guide
    ws.column_dimensions['Q'].width = 15   # Spacer
    ws.column_dimensions['R'].width = 15  # payment status
    ws.column_dimensions['S'].width = 15  # supplier / guide
    ws.column_dimensions['T'].width = 18  # PROFIT PER PRODUCT
    ws.column_dimensions['U'].width = 12  # PROFIT on file

    # Populate sum formula at L5
    # Sum of Column M (13) from data_start_row to last row
    last_data_row = current_row - 1
    if last_data_row >= data_start_row:
        cell = ws.cell(row=5, column=12, value=f"=SUM(M{data_start_row}:M{last_data_row})")
        cell.number_format = '#,##0.00'

        # Populate sum of Profit (Column T) at U4
        cell = ws.cell(row=5, column=21, value=f"=SUM(T{data_start_row}:T{last_data_row})")
        cell.number_format = '#,##0.00'

    # Format up to row 500
    # Ensure center alignment for A, B, C, D, G, H, I, J and number format for M
    center_align_style = Alignment(horizontal='center', vertical='center')

    for r in range(1, 501):
        # Column M: number format
        ws.cell(row=r, column=13).number_format = '#,##0.00'

        # Column C(3), D(4): date format
        ws.cell(row=r, column=3).number_format = 'dd-mmm-yy'
        ws.cell(row=r, column=4).number_format = 'dd-mmm-yy'

        # Center align specific columns: A(1), B(2), C(3), D(4), E(5), G(7), H(8), I(9), J(10)
        for col in [1, 2, 3, 4, 5, 7, 8, 9, 10]:
            # Skip C4 (row 4, column 3)
            if r == 4 and col == 3:
                continue
            ws.cell(row=r, column=col).alignment = center_align_style

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"inputsheet_export_{package.package_reference}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# ──────────────────────────────────────────────────────────────
# Invoice & Supplier Expense views
# ──────────────────────────────────────────────────────────────

from functools import wraps
from django.core.exceptions import PermissionDenied


def superuser_or_owner_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if request.user.is_superuser or request.user.groups.filter(name='owner').exists():
            return view_func(request, *args, **kwargs)
        raise PermissionDenied
    return _wrapped

def get_grouped_tour_data(package):
    """
    Extract tour data grouped by hotel stay, following the same logic as export_tourday_excel.
    Returns (final_items, hotel_groups_list) where:
    - final_items: list of dicts with arrival_date, departure_date, nights, service_name, price, etc.
    - hotel_groups_list: list of hotel group dicts
    """
    from datetime import timedelta
    import re

    SERVICE_TYPE_ORDER = {
        'transfer': 1,
        'package': 3,
        'tour': 4,
        'custom': 5,
        'zero': 6,
    }

    def get_service_type_order(service_type_name, service_name=''):
        """Get sort order for service type. Special case: ** in name acts as transfer."""
        if '**' in service_name:
            return 1
        name_lower = service_type_name.lower() if service_type_name else ''
        for key, order in SERVICE_TYPE_ORDER.items():
            if key in name_lower:
                return order
        return 99

    tour_days = list(package.tour_days.all().order_by('date'))

    # Build hotel info lookup from hotel_costs
    hotel_info_lookup = {}
    for cost in (package.hotel_costs or []):
        hotel_name = cost.get('name', '')
        room_count = cost.get('room', 1)
        room_type = cost.get('type', '')
        extra_bed_price = cost.get('extraBedPrice')
        room_price = cost.get('price', 0)
        nights = cost.get('nights', 1)
        promotion = cost.get('promotion', '')
        date_str = cost.get('date', '')

        # Calculate total price
        extra_bed = float(extra_bed_price) if extra_bed_price and float(extra_bed_price) > 0 else 0
        room_price_val = float(room_price) if room_price else 0
        nights_val = int(nights) if nights else 1
        rooms_val = int(room_count) if room_count else 1

        total_price = (room_price_val * rooms_val + extra_bed) * nights_val

        # Build display name
        room_part = ""
        if room_count and room_type:
            room_part = f"{room_count}x {room_type}"
        elif room_type:
            room_part = room_type

        if extra_bed > 0:
            display_name = f"{room_part} + extra bed" if room_part else ""
        else:
            display_name = room_part

        if hotel_name not in hotel_info_lookup:
            hotel_info_lookup[hotel_name] = []
        hotel_info_lookup[hotel_name].append({
            'display': display_name,
            'price': total_price,
            'promotion': promotion,
            'date_str': date_str,
            'nights': nights_val,
            'room_count': rooms_val,
            'room_price': room_price_val,
            'extra_bed': extra_bed,
        })

    # Group tour days by hotel
    hotel_groups_list = []
    last_hotel_name = None
    current_group = None

    for day in tour_days:
        hotel_name = day.hotel.name if day.hotel else 'No Hotel'

        if hotel_name != last_hotel_name:
            current_group = {
                'hotel_name': hotel_name,
                'arrival_date': day.date,
                'departure_date': day.date + timedelta(days=1),
                'nights': 1,
                'services': [],
                'guide_services': [],
                'dates': [day.date],
            }
            hotel_groups_list.append(current_group)
            last_hotel_name = hotel_name
        else:
            current_group['departure_date'] = day.date + timedelta(days=1)
            current_group['nights'] += 1
            current_group['dates'].append(day.date)

        # Collect services for this day
        for service in day.services.all():
            service_type_name = service.service.service_type.name if service.service.service_type else ''
            service_name = service.service.name
            # Try to get supplier name
            supplier_name = ''
            if hasattr(service.service, 'supplier') and service.service.supplier:
                supplier_name = service.service.supplier.name
            else:
                supplier_name = service_name

            current_group['services'].append({
                'date': day.date,
                'service_type': service_type_name,
                'sort_order': get_service_type_order(service_type_name, service_name),
                'name': service_name,
                'price': service.price_at_booking,
                'supplier': supplier_name,
                'service_id': service.service.id,
            })

        # Collect guide services
        for guide_service in day.guide_services.all():
            current_group['guide_services'].append({
                'date': day.date,
                'service_type': 'custom',
                'sort_order': 5,
                'name': guide_service.guide_service.name,
                'price': guide_service.price_at_booking,
                'supplier': guide_service.guide_service.name,
                'service_id': None,
            })

    # Build final items grouped by hotel
    final_items = []

    for hotel_data in hotel_groups_list:
        hotel_name = hotel_data['hotel_name']

        # Combine regular services and guide services, sort by date
        all_services = hotel_data['services'] + hotel_data['guide_services']
        all_services.sort(key=lambda x: x['date'])

        # Build hotel rows
        hotel_rows = []
        if 'No Hotel' not in hotel_name:
            matching_configs = []
            if hotel_name in hotel_info_lookup:
                room_configs = hotel_info_lookup[hotel_name]
                group_dates = hotel_data.get('dates', [hotel_data['arrival_date']])

                for config in room_configs:
                    c_date = config.get('date_str', '')
                    if not c_date:
                        continue

                    # Check if config date matches any day in the group
                    matched_date = None
                    for d in group_dates:
                        day_str_pad = f"{d.day:02d}"
                        month_str = d.strftime('%b').lower()

                        if (c_date.startswith(day_str_pad) or c_date.startswith(str(d.day))) and month_str in c_date.lower():
                            matched_date = d
                            break

                    if matched_date:
                        config_copy = config.copy()
                        config_copy['specific_arrival_date'] = matched_date
                        matching_configs.append(config_copy)

                # Fallback if no specific date match found
                if not matching_configs:
                    matching_configs.append({
                        'display': '',
                        'price': 0,
                        'promotion': '',
                        'nights': hotel_data['nights'],
                    })

                # Sort by date
                matching_configs.sort(key=lambda x: x.get('specific_arrival_date') or hotel_data['arrival_date'])
            else:
                # No room config found, use defaults
                matching_configs.append({
                    'display': '',
                    'price': 0,
                    'promotion': '',
                    'nights': hotel_data['nights'],
                })

            # Create hotel rows for each room config
            for room_config in matching_configs:
                hotel_display_name = hotel_name
                if room_config.get('display'):
                    hotel_display_name = f"{hotel_name}, {room_config['display']}"

                arrival_date = room_config.get('specific_arrival_date', hotel_data['arrival_date'])
                nights = room_config.get('nights', hotel_data['nights'])

                try:
                    nights_int = int(nights) if nights else 0
                    departure_date = arrival_date + timedelta(days=nights_int)
                except:
                    departure_date = hotel_data['departure_date']

                hotel_rows.append({
                    'arrival_date': arrival_date,
                    'departure_date': departure_date,
                    'nights': nights,
                    'service_name': hotel_display_name,
                    'price': room_config.get('price', 0),
                    'promotion': room_config.get('promotion', ''),
                    'is_hotel': True,
                    'supplier': hotel_name,
                    'room_price': room_config.get('room_price', 0),
                    'room_count': room_config.get('room_count', 1),
                    'extra_bed_price': room_config.get('extra_bed', 0),
                })

        # Determine which services should come before hotel
        def should_be_before_hotel(svc):
            service_type = svc.get('service_type', '').lower()
            service_name = svc.get('name', '')
            return 'transfer' in service_type or '**' in service_name or 'transfer' in service_name.lower()

        last_before_hotel_idx = -1
        for idx, svc in enumerate(all_services):
            if should_be_before_hotel(svc):
                last_before_hotel_idx = idx

        # Build final items in correct order
        if last_before_hotel_idx >= 0:
            # Transfers/transfer services before hotel
            for svc in all_services[:last_before_hotel_idx + 1]:
                final_items.append({
                    'arrival_date': svc['date'],
                    'departure_date': None,
                    'nights': None,
                    'service_name': svc['name'],
                    'price': float(svc['price']) if svc['price'] else 0,
                    'is_hotel': False,
                    'supplier': svc.get('supplier', svc['name']),
                    'service_id': svc.get('service_id'),
                })
            # Hotel rows
            for hotel_row in hotel_rows:
                final_items.append(hotel_row)
            # Remaining services after hotel
            for svc in all_services[last_before_hotel_idx + 1:]:
                final_items.append({
                    'arrival_date': svc['date'],
                    'departure_date': None,
                    'nights': None,
                    'service_name': svc['name'],
                    'price': float(svc['price']) if svc['price'] else 0,
                    'is_hotel': False,
                    'supplier': svc.get('supplier', svc['name']),
                    'service_id': svc.get('service_id'),
                })
        else:
            # No transfers - hotel first, then services
            for hotel_row in hotel_rows:
                final_items.append(hotel_row)
            for svc in all_services:
                final_items.append({
                    'arrival_date': svc['date'],
                    'departure_date': None,
                    'nights': None,
                    'service_name': svc['name'],
                    'price': float(svc['price']) if svc['price'] else 0,
                    'is_hotel': False,
                    'supplier': svc.get('supplier', svc['name']),
                    'service_id': svc.get('service_id'),
                })

    # Add extra costs at the end
    extra_costs = package.extra_costs or []
    for extra_cost in extra_costs:
        cost_name = extra_cost.get('item', '') or extra_cost.get('name', '')
        cost_amount = extra_cost.get('amount', 0)
        if cost_name:
            final_items.append({
                'arrival_date': None,
                'departure_date': None,
                'nights': None,
                'service_name': cost_name,
                'price': float(cost_amount) if cost_amount else 0,
                'is_hotel': False,
                'supplier': '',
                'is_extra_cost': True,
            })

    # Add discounts at the end
    discounts = package.discounts or []
    for discount in discounts:
        discount_name = discount.get('item', '') or discount.get('name', '')
        discount_amount = discount.get('amount', 0)
        if discount_name:
            final_items.append({
                'arrival_date': None,
                'departure_date': None,
                'nights': None,
                'service_name': discount_name,
                'price': -abs(float(discount_amount)) if discount_amount else 0,
                'is_hotel': False,
                'supplier': '',
                'is_discount': True,
            })

    return final_items, hotel_groups_list


def build_prepopulated_invoice_data(package):
    """
    Returns (invoice_items, supplier_expenses) as lists of dicts
    pre-populated from the TourPackageQuote data, following the same
    grouping and ordering as export_tourday_excel.
    """
    from .models import ServiceExpenseTemplate, ServicePrice

    grouped_items, _ = get_grouped_tour_data(package)
    invoice_items = []
    supplier_expenses = []

    # Build service_id → service_price_id lookup using the package's tour_pack_type
    service_ids = [item['service_id'] for item in grouped_items if item.get('service_id')]
    tour_pack_type_id = package.tour_pack_type_id if package.tour_pack_type_id else None

    service_price_by_service = {}
    if service_ids and tour_pack_type_id:
        for sp in ServicePrice.objects.filter(
            service_id__in=service_ids, tour_pack_type_id=tour_pack_type_id
        ):
            service_price_by_service[sp.service_id] = sp.id

    # Prefetch expense templates keyed by service_price_id
    service_price_ids = list(service_price_by_service.values())
    templates_qs = ServiceExpenseTemplate.objects.filter(
        service_price_id__in=service_price_ids
    ).select_related('supplier', 'supplier_service')
    templates_by_service_price = {}
    for tmpl in templates_qs:
        templates_by_service_price.setdefault(tmpl.service_price_id, []).append(tmpl)

    for i, item in enumerate(grouped_items):
        # Handle discounts
        if item.get('is_discount'):
            invoice_items.append({
                'description': item['service_name'],
                'quantity': '1',
                'unit_price': str(item['price']),
                'amount': str(item['price']),
                'item_type': 'Discount',
                'order': i,
            })
            continue

        # Build description
        desc_parts = [item['service_name']]
        if item.get('arrival_date') and not item.get('is_hotel'):
            desc_parts.append(f"({item['arrival_date']})")
        elif item.get('is_hotel'):
            arrival = item['arrival_date'].strftime('%d-%b') if item['arrival_date'] else ''
            departure = item['departure_date'].strftime('%d-%b') if item['departure_date'] else ''
            if arrival and departure:
                desc_parts.append(f"- {arrival} to {departure}, {item['nights']} nights")
            if item.get('promotion'):
                desc_parts.append(f"({item['promotion']})")

        desc = ' '.join(filter(None, desc_parts))
        amt = str(abs(Decimal(str(item['price'] or 0))))

        # Determine item type
        if item.get('is_extra_cost'):
            item_type = 'Extra'
        elif item.get('is_hotel'):
            item_type = 'Hotel'
        else:
            item_type = 'Service'

        invoice_items.append({
            'description': desc,
            'quantity': '1',
            'unit_price': amt,
            'amount': amt,
            'item_type': item_type,
            'order': i,
        })

        # Build supplier expenses: use templates if defined, otherwise fall back to single derived expense
        service_id = item.get('service_id')
        service_price_id = service_price_by_service.get(service_id) if service_id else None
        templates = templates_by_service_price.get(service_price_id, []) if service_price_id else []

        if templates:
            for tmpl in templates:
                sname = tmpl.supplier_name or (tmpl.supplier.name if tmpl.supplier else '')
                supplier_expenses.append({
                    'supplier_name': sname,
                    'supplier_id': tmpl.supplier_id,
                    'supplier_service_id': tmpl.supplier_service_id,
                    'description': (tmpl.supplier_service.name if tmpl.supplier_service else '') or desc,
                    'unit_price': str(tmpl.unit_price),
                    'amount': str(tmpl.unit_price),
                    'due_date': '',
                    'status': 'Pending',
                    'reference_number': '',
                    'order': len(supplier_expenses),
                    'source_item_index': i,
                })
        else:
            is_hotel = item.get('is_hotel')
            supplier_name = item.get('supplier', '')
            if not supplier_name and is_hotel:
                supplier_name = item['service_name'].split(',')[0] if ',' in item['service_name'] else item['service_name']
            # Hotels have no cost template — default to 0 so staff enter actual supplier cost.
            default_amount = '0' if is_hotel else amt
            expense = {
                'supplier_name': supplier_name,
                'description': desc,
                'unit_price': default_amount,
                'amount': default_amount,
                'due_date': '',
                'status': 'Pending',
                'reference_number': '',
                'order': len(supplier_expenses),
                'source_item_index': i,
            }
            if is_hotel:
                expense.update({
                    'room_count': item.get('room_count', 1),
                    'nights': item.get('nights', 1),
                    'room_price': 0,
                    'extra_bed_price': 0,
                    'promotion': item.get('promotion', ''),
                })
            supplier_expenses.append(expense)

    return invoice_items, supplier_expenses


def _parse_item_meta(description):
    """Parse InvoiceItem description encoded as:
    service_name|||arrival|||departure|||nights|||room_count|||room_price|||extra_bed_price
    Returns a dict; all date/numeric fields are None when absent.
    """
    parts = description.split('|||')

    def _date(s):
        if not s or not s.strip():
            return None
        try:
            return datetime.strptime(s.strip(), '%Y-%m-%d').date()
        except ValueError:
            return None

    def _num(s, default=None):
        if s is None or str(s).strip() == '':
            return default
        try:
            return float(str(s).strip())
        except (ValueError, TypeError):
            return default

    return {
        'service_name':    parts[0] if parts else description,
        'arrival_date':    _date(parts[1]) if len(parts) > 1 else None,
        'departure_date':  _date(parts[2]) if len(parts) > 2 else None,
        'nights':          parts[3].strip() if len(parts) > 3 and parts[3].strip() else None,
        'room_count':      _num(parts[4], None) if len(parts) > 4 else None,
        'room_price':      _num(parts[5], None) if len(parts) > 5 else None,
        'extra_bed_price': _num(parts[6], None) if len(parts) > 6 else None,
        'promotion':       parts[7].strip() if len(parts) > 7 and parts[7].strip() else None,
    }


@login_required
def invoice_service_search(request):
    q = request.GET.get('q', '').strip()
    tour_pack_type_id = request.GET.get('tour_pack_type')
    if not tour_pack_type_id or not q:
        return JsonResponse([], safe=False)

    prices = ServicePrice.objects.filter(
        tour_pack_type_id=tour_pack_type_id,
        service__name__icontains=q,
    ).select_related('service').prefetch_related(
        'expense_templates__supplier',
        'expense_templates__supplier_service',
    )[:20]

    results = []
    for sp in prices:
        results.append({
            'service_id': sp.service_id,
            'name': sp.service.name,
            'price': str(sp.price),
            'expense_templates': [
                {
                    'supplier_id': t.supplier_id,
                    'supplier_name': t.supplier_name or (t.supplier.name if t.supplier else ''),
                    'description': t.supplier_service.name if t.supplier_service else '',
                    'unit_price': str(t.unit_price),
                    'supplier_service_id': t.supplier_service_id,
                }
                for t in sp.expense_templates.all()
            ],
        })
    return JsonResponse(results, safe=False)


def _get_suppliers_data():
    return [
        {
            'id': s.id,
            'name': s.name,
            'services': [
                {'id': ss.id, 'name': ss.name, 'cost': str(ss.cost)}
                for ss in s.supplier_services.all()
            ],
        }
        for s in Supplier.objects.prefetch_related('supplier_services').all()
    ]


@login_required
@superuser_or_owner_required
def create_invoice(request, package_reference):
    package = get_object_or_404(TourPackageQuote, package_reference=package_reference)
    agencies = Agency.objects.all()

    if request.method == 'POST':
        agency_id = request.POST.get('agency') or None
        issue_date = parse_custom_date(request.POST.get('issue_date'))
        due_date = parse_custom_date(request.POST.get('due_date'))
        notes = request.POST.get('notes', '')
        status = request.POST.get('status', Invoice.STATUS_DRAFT)
        items_json = request.POST.get('invoice_items_json', '[]')
        expenses_json = request.POST.get('supplier_expenses_json', '[]')

        try:
            items_data = json.loads(items_json)
            expenses_data = json.loads(expenses_json)
        except json.JSONDecodeError:
            items_data, expenses_data = [], []

        with transaction.atomic():
            invoice = Invoice.objects.create(
                tour_package=package,
                tour_pack_type=package.tour_pack_type,
                agency_id=agency_id,
                issue_date=issue_date,
                due_date=due_date,
                notes=notes,
                status=status,
                created_by=request.user,
            )

            for i, item in enumerate(items_data):
                qty = safe_decimal(item.get('quantity', 1), Decimal('1'))
                unit_price = safe_decimal(item.get('unit_price', 0))
                InvoiceItem.objects.create(
                    invoice=invoice,
                    description=item.get('description', '')[:500],
                    quantity=qty,
                    unit_price=unit_price,
                    amount=qty * unit_price,
                    item_type=item.get('item_type', 'Other'),
                    order=i,
                )

            supplier_cache = {s.name: s for s in Supplier.objects.all()}
            service_id_set = {
                int(exp['supplier_service_id'])
                for exp in expenses_data if exp.get('supplier_service_id')
            }
            supplier_service_cache = (
                {ss.id: ss for ss in SupplierService.objects.filter(id__in=service_id_set)}
                if service_id_set else {}
            )
            for i, exp in enumerate(expenses_data):
                sname = exp.get('supplier_name', '')[:200]
                src_idx = exp.get('source_item_index')
                ss_id = exp.get('supplier_service_id')
                SupplierExpense.objects.create(
                    invoice=invoice,
                    supplier=supplier_cache.get(sname),
                    supplier_service=supplier_service_cache.get(int(ss_id)) if ss_id else None,
                    supplier_name=sname,
                    description=exp.get('description', '')[:500],
                    unit_price=safe_decimal(exp.get('unit_price', 0)),
                    amount=safe_decimal(exp.get('amount', 0)),
                    due_date=parse_custom_date(exp.get('due_date')),
                    status=exp.get('status', 'Pending'),
                    reference_number=exp.get('reference_number', '')[:100],
                    order=i,
                    source_item_index=int(src_idx) if src_idx is not None else None,
                )

            invoice.recalculate_total()

        messages.success(request, f'Invoice {invoice.invoice_number} created successfully.')
        return redirect('invoice_detail', invoice_id=invoice.id)

    invoice_items, supplier_expenses = build_prepopulated_invoice_data(package)
    grouped_items, _ = get_grouped_tour_data(package)
    today = datetime.now().date()

    context = {
        'package': package,
        'agencies': agencies,
        'invoice_items': invoice_items,
        'supplier_expenses': supplier_expenses,
        'grouped_items': grouped_items,
        'suppliers_data': _get_suppliers_data(),
        'invoice_status_choices': Invoice.STATUS_CHOICES,
        'invoice_item_types': InvoiceItem.ITEM_TYPE_CHOICES,
        'today': today,
        'tour_pack_type_id': package.tour_pack_type_id or '',
    }
    return render(request, 'tour_quote/create_invoice.html', context)


@login_required
@superuser_or_owner_required
def edit_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    agencies = Agency.objects.all()

    if request.method == 'POST':
        agency_id = request.POST.get('agency') or None
        issue_date = parse_custom_date(request.POST.get('issue_date'))
        due_date = parse_custom_date(request.POST.get('due_date'))
        notes = request.POST.get('notes', '')
        status = request.POST.get('status', Invoice.STATUS_DRAFT)
        items_json = request.POST.get('invoice_items_json', '[]')
        expenses_json = request.POST.get('supplier_expenses_json', '[]')

        try:
            items_data = json.loads(items_json)
            expenses_data = json.loads(expenses_json)
        except json.JSONDecodeError:
            items_data, expenses_data = [], []

        with transaction.atomic():
            invoice.agency_id = agency_id
            invoice.issue_date = issue_date
            invoice.due_date = due_date
            invoice.notes = notes
            invoice.status = status
            invoice.save()

            invoice.items.all().delete()
            invoice.supplier_expenses.all().delete()

            for i, item in enumerate(items_data):
                qty = safe_decimal(item.get('quantity', 1), Decimal('1'))
                unit_price = safe_decimal(item.get('unit_price', 0))
                InvoiceItem.objects.create(
                    invoice=invoice,
                    description=item.get('description', '')[:500],
                    quantity=qty,
                    unit_price=unit_price,
                    amount=qty * unit_price,
                    item_type=item.get('item_type', 'Other'),
                    order=i,
                )

            supplier_cache = {s.name: s for s in Supplier.objects.all()}
            service_id_set = {
                int(exp['supplier_service_id'])
                for exp in expenses_data if exp.get('supplier_service_id')
            }
            supplier_service_cache = (
                {ss.id: ss for ss in SupplierService.objects.filter(id__in=service_id_set)}
                if service_id_set else {}
            )
            for i, exp in enumerate(expenses_data):
                sname = exp.get('supplier_name', '')[:200]
                src_idx = exp.get('source_item_index')
                ss_id = exp.get('supplier_service_id')
                SupplierExpense.objects.create(
                    invoice=invoice,
                    supplier=supplier_cache.get(sname),
                    supplier_service=supplier_service_cache.get(int(ss_id)) if ss_id else None,
                    supplier_name=sname,
                    description=exp.get('description', '')[:500],
                    unit_price=safe_decimal(exp.get('unit_price', 0)),
                    amount=safe_decimal(exp.get('amount', 0)),
                    due_date=parse_custom_date(exp.get('due_date')),
                    status=exp.get('status', 'Pending'),
                    reference_number=exp.get('reference_number', '')[:100],
                    order=i,
                    source_item_index=int(src_idx) if src_idx is not None else None,
                )

            invoice.recalculate_total()

        messages.success(request, f'Invoice {invoice.invoice_number} updated.')
        return redirect('invoice_detail', invoice_id=invoice.id)

    existing_items = list(invoice.items.all().order_by('order'))
    package_items, _ = get_grouped_tour_data(invoice.tour_package)

    grouped_items = []
    for i in range(max(len(package_items), len(existing_items))):
        pkg = package_items[i] if i < len(package_items) else None
        inv = existing_items[i] if i < len(existing_items) else None

        if inv:
            saved = _parse_item_meta(inv.description)

            def _pkg_date_str(key):
                val = pkg.get(key, '') if pkg else ''
                return val.strftime('%Y-%m-%d') if hasattr(val, 'strftime') else str(val or '')

            arr = (saved['arrival_date'].strftime('%Y-%m-%d') if saved['arrival_date'] else '') or _pkg_date_str('arrival_date')
            dep = (saved['departure_date'].strftime('%Y-%m-%d') if saved['departure_date'] else '') or _pkg_date_str('departure_date')

            grouped_items.append({
                'arrival_date':    arr,
                'departure_date':  dep,
                'nights':          saved['nights'] or (str(pkg.get('nights', '')) if pkg else ''),
                'service_name':    saved['service_name'],
                'price':           str(inv.amount),
                'is_hotel':        inv.item_type == InvoiceItem.ITEM_TYPE_HOTEL,
                'is_discount':     inv.item_type == InvoiceItem.ITEM_TYPE_DISCOUNT,
                'is_extra_cost':   inv.item_type == InvoiceItem.ITEM_TYPE_EXTRA,
                'room_price':      saved['room_price']      if saved['room_price']      is not None else (pkg.get('room_price', 0)      if pkg else 0),
                'room_count':      saved['room_count']      if saved['room_count']      is not None else (pkg.get('room_count', 1)      if pkg else 1),
                'extra_bed_price': saved['extra_bed_price'] if saved['extra_bed_price'] is not None else (pkg.get('extra_bed_price', 0) if pkg else 0),
                'promotion':       saved['promotion']       or (pkg.get('promotion', '') if pkg else ''),
            })
        elif pkg:
            # Package item not saved to invoice (shouldn't happen, but handle)
            grouped_items.append({
                'arrival_date': pkg.get('arrival_date', ''),
                'departure_date': pkg.get('departure_date', ''),
                'nights': pkg.get('nights', ''),
                'service_name': pkg.get('service_name', ''),
                'price': str(pkg.get('price', 0)),
                'is_hotel': pkg.get('is_hotel', False),
                'is_discount': pkg.get('is_discount', False),
                'is_extra_cost': pkg.get('is_extra_cost', False),
                'promotion': pkg.get('promotion', ''),
                'room_price': pkg.get('room_price', 0),
                'room_count': pkg.get('room_count', 1),
                'extra_bed_price': pkg.get('extra_bed_price', 0),
            })

    existing_expenses = []
    for exp in invoice.supplier_expenses.all().order_by('order'):
        exp_dict = {
            'supplier_name': exp.supplier_name,
            'supplier_id': exp.supplier_id,
            'supplier_service_id': exp.supplier_service_id,
            'description': exp.description,
            'unit_price': str(exp.unit_price),
            'amount': str(exp.amount),
            'due_date': exp.due_date.strftime('%d-%b-%y') if exp.due_date else '',
            'status': exp.status,
            'reference_number': exp.reference_number or '',
            'order': exp.order,
            'source_item_index': exp.source_item_index,
        }
        # Reconstruct hotel fields from linked item so the expense panel can show/edit them
        if exp.source_item_index is not None and exp.source_item_index < len(grouped_items):
            item = grouped_items[exp.source_item_index]
            if item.get('is_hotel'):
                room_count = int(item.get('room_count', 1) or 1)
                nights = int(item.get('nights', 1) or 1)
                amt = float(exp.amount or 0)
                # Derive room_price from saved amount (assume no extra bed in saved data)
                room_price = (amt / (room_count * nights)) if room_count * nights > 0 else 0
                exp_dict.update({
                    'room_count': room_count,
                    'nights': nights,
                    'room_price': room_price,
                    'extra_bed_price': 0,
                    'promotion': item.get('promotion', ''),
                })
        existing_expenses.append(exp_dict)

    context = {
        'invoice': invoice,
        'agencies': agencies,
        'grouped_items': grouped_items,
        'supplier_expenses': existing_expenses,
        'suppliers_data': _get_suppliers_data(),
        'invoice_status_choices': Invoice.STATUS_CHOICES,
        'tour_pack_type_id': (invoice.tour_pack_type_id or invoice.tour_package.tour_pack_type_id or ''),
    }
    return render(request, 'tour_quote/edit_invoice.html', context)


@login_required
@superuser_or_owner_required
def export_service_expense_template(request, tour_pack_type_id):
    from .admin import ServiceExpenseTemplateResource
    from import_export.formats.base_formats import XLSX

    tour_pack_type = get_object_or_404(TourPackType, id=tour_pack_type_id)
    queryset = ServiceExpenseTemplate.objects.filter(
        service_price__tour_pack_type=tour_pack_type
    ).select_related('service_price__service', 'service_price__tour_pack_type', 'supplier', 'supplier_service')

    resource = ServiceExpenseTemplateResource()
    dataset = resource.export(queryset=queryset)
    format = XLSX()
    export_data = format.export_data(dataset)

    response = HttpResponse(
        export_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"service_expense_template_{tour_pack_type.name}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@superuser_or_owner_required
def invoice_detail(request, invoice_id):
    # per-service profit mapping with grouping visuals
    invoice = get_object_or_404(Invoice, id=invoice_id)
    items = list(invoice.items.all().order_by('order'))
    expenses = list(invoice.supplier_expenses.all().order_by('order'))
    total_expenses = sum(e.amount for e in expenses)
    margin = invoice.total_amount - total_expenses

    grouped_items, _ = get_grouped_tour_data(invoice.tour_package)

    # Group expenses by source_item_index and calculate per-service profit
    expenses_by_item = {}
    for exp in expenses:
        idx = exp.source_item_index
        if idx is not None:
            expenses_by_item.setdefault(idx, []).append(exp)

    service_profits = {}
    for idx, item in enumerate(items):
        item_expenses = expenses_by_item.get(idx, [])
        total_exp = sum(e.amount for e in item_expenses)
        service_profits[idx] = {
            'description': item.description,
            'selling_price': item.amount,
            'profit': item.amount - total_exp,
        }

    service_expense_counts = {idx: len(exps) for idx, exps in expenses_by_item.items()}
    # total expenses per service group for first-row display
    group_totals = {idx: sum(e.amount for e in exps) for idx, exps in expenses_by_item.items()}

    expenses_with_service = []
    seen_indices = set()
    for exp in expenses:
        idx = exp.source_item_index
        sp = service_profits.get(idx)
        entry = {
            'expense': exp,
            'service_description': sp['description'] if sp else None,
            'selling_price': sp['selling_price'] if sp else None,
            'profit': sp['profit'] if sp else None,
            'is_first_for_service': idx is not None and idx not in seen_indices,
            'service_expense_count': service_expense_counts.get(idx, 0),
            'group_total': group_totals.get(idx, 0),
        }
        # Attach hotel metadata for display (from saved invoice item, fallback to package)
        if idx is not None and idx < len(items):
            item_obj = items[idx]
            if item_obj.item_type == InvoiceItem.ITEM_TYPE_HOTEL:
                saved = _parse_item_meta(item_obj.description)
                gi = grouped_items[idx] if idx < len(grouped_items) else None
                entry['is_hotel'] = True
                entry['room_count'] = saved['room_count'] if saved['room_count'] is not None else (gi.get('room_count', 1) if gi else 1)
                entry['nights'] = saved['nights'] or (gi.get('nights', '') if gi else '')
                entry['room_price'] = saved['room_price'] if saved['room_price'] is not None else (gi.get('room_price', 0) if gi else 0)
                entry['extra_bed_price'] = saved['extra_bed_price'] if saved['extra_bed_price'] is not None else (gi.get('extra_bed_price', 0) if gi else 0)
                entry['promotion'] = saved['promotion'] or (gi.get('promotion', '') if gi else '')
                entry['arrival_date'] = saved['arrival_date'] or (gi.get('arrival_date') if gi else None)
                entry['departure_date'] = saved['departure_date'] or (gi.get('departure_date') if gi else None)
                # Derive actual cost per night from the expense amount for display in Description column
                rc = int(entry['room_count'] or 1)
                nt = int(entry['nights'] or 1)
                amt = float(exp.amount or 0)
                entry['expense_room_price'] = (amt / (rc * nt)) if rc * nt > 0 else 0
        expenses_with_service.append(entry)
        if idx is not None:
            seen_indices.add(idx)

    # Dates and hotel pricing come from saved description metadata,
    # falling back to current package data only when absent.
    items_with_meta = []
    for i, item in enumerate(items):
        saved = _parse_item_meta(item.description)
        gi = grouped_items[i] if i < len(grouped_items) else None

        arrival   = saved['arrival_date']   or (gi.get('arrival_date')   if gi else None)
        departure = saved['departure_date'] or (gi.get('departure_date') if gi else None)
        nights      = saved['nights']          or (str(gi.get('nights', ''))      if gi else '')
        room_price  = saved['room_price']      if saved['room_price']      is not None else (gi.get('room_price', 0)      if gi else 0)
        room_count  = saved['room_count']      if saved['room_count']      is not None else (gi.get('room_count', 1)      if gi else 1)
        extra_bed   = saved['extra_bed_price'] if saved['extra_bed_price'] is not None else (gi.get('extra_bed_price', 0) if gi else 0)
        promotion   = saved['promotion']       or (gi.get('promotion', '')          if gi else '')

        meta = {'arrival_date': arrival, 'departure_date': departure}
        if item.item_type == InvoiceItem.ITEM_TYPE_HOTEL and room_price:
            meta.update({
                'room_price':      room_price,
                'room_count':      int(room_count),
                'nights':          nights,
                'extra_bed_price': extra_bed,
                'promotion':       promotion,
            })
        items_with_meta.append({'item': item, 'meta': meta})

    context = {
        'invoice': invoice,
        'items_with_meta': items_with_meta,
        'expenses_with_service': expenses_with_service,
        'expenses': expenses,
        'total_expenses': total_expenses,
        'margin': margin,
    }
    return render(request, 'tour_quote/invoice_detail.html', context)


@login_required
@superuser_or_owner_required
def invoice_list(request):
    invoices = (
        Invoice.objects
        .select_related('tour_package', 'agency')
        .annotate(
            expense_total=Coalesce(Sum('supplier_expenses__amount'), Decimal('0.00')),
        )
        .annotate(
            margin=ExpressionWrapper(
                F('total_amount') - F('expense_total'),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            )
        )
        .order_by('-created_at')
    )
    status_filter = request.GET.get('status', '')
    if status_filter:
        invoices = invoices.filter(status=status_filter)

    paginator = Paginator(invoices, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'status_choices': Invoice.STATUS_CHOICES,
        'current_status': status_filter,
    }
    return render(request, 'tour_quote/invoice_list.html', context)


@login_required
@superuser_or_owner_required
def invoice_pdf(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    items = list(invoice.items.all().order_by('order'))

    # Dates and hotel pricing come from saved description metadata,
    # falling back to current package data only when absent.
    grouped_items, _ = get_grouped_tour_data(invoice.tour_package)
    items_with_meta = []
    for i, item in enumerate(items):
        saved = _parse_item_meta(item.description)
        gi = grouped_items[i] if i < len(grouped_items) else None

        arrival   = saved['arrival_date']   or (gi.get('arrival_date')   if gi else None)
        departure = saved['departure_date'] or (gi.get('departure_date') if gi else None)
        nights      = saved['nights']          or (str(gi.get('nights', ''))      if gi else '')
        room_price  = saved['room_price']      if saved['room_price']      is not None else (gi.get('room_price', 0)      if gi else 0)
        room_count  = saved['room_count']      if saved['room_count']      is not None else (gi.get('room_count', 1)      if gi else 1)
        extra_bed   = saved['extra_bed_price'] if saved['extra_bed_price'] is not None else (gi.get('extra_bed_price', 0) if gi else 0)
        promotion   = saved['promotion']       or (gi.get('promotion', '')          if gi else '')

        meta = {'arrival_date': arrival, 'departure_date': departure}
        if item.item_type == InvoiceItem.ITEM_TYPE_HOTEL and room_price:
            meta.update({
                'room_price':      room_price,
                'room_count':      int(room_count),
                'nights':          nights,
                'extra_bed_price': extra_bed,
                'promotion':       promotion,
            })
        items_with_meta.append({'item': item, 'meta': meta})

    logo_data_uri = None
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'image', 'rsz_animo1.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_data_uri = f'data:image/png;base64,{base64.b64encode(f.read()).decode()}'

    html_string = render_to_string('tour_quote/invoice_pdf.html', {
        'invoice': invoice,
        'items': items,
        'items_with_meta': items_with_meta,
        'logo_data_uri': logo_data_uri,
    })

    response = HttpResponse(content_type='application/pdf')
    safe_num = ''.join(c for c in (invoice.invoice_number or 'invoice') if c.isalnum() or c in '-_')
    response['Content-Disposition'] = f'inline; filename="Invoice_{safe_num}.pdf"'
    HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(
        response,
        stylesheets=[CSS(string='body { font-family: sans-serif; font-size: 12px; }')]
    )
    return response


@login_required
@superuser_or_owner_required
def payment_list_view(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    expenses = list(invoice.supplier_expenses.all().order_by('order'))
    total_expenses = sum(e.amount for e in expenses)

    if request.GET.get('pdf'):
        logo_data_uri = None
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'image', 'rsz_animo1.png')
        if os.path.exists(logo_path):
            with open(logo_path, 'rb') as f:
                logo_data_uri = f'data:image/png;base64,{base64.b64encode(f.read()).decode()}'

        html_string = render_to_string('tour_quote/payment_list_pdf.html', {
            'invoice': invoice,
            'expenses': expenses,
            'total_expenses': total_expenses,
            'logo_data_uri': logo_data_uri,
        })
        response = HttpResponse(content_type='application/pdf')
        safe_num = ''.join(c for c in (invoice.invoice_number or 'invoice') if c.isalnum() or c in '-_')
        response['Content-Disposition'] = f'inline; filename="PaymentList_{safe_num}.pdf"'
        HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(
            response,
            stylesheets=[CSS(string='@page { size: A4; margin: 2cm; } body { font-family: sans-serif; font-size: 12px; }')]
        )
        return response

    context = {
        'invoice': invoice,
        'expenses': expenses,
        'total_expenses': total_expenses,
    }
    return render(request, 'tour_quote/payment_list.html', context)


@login_required
@superuser_or_owner_required
def update_supplier_expense_status(request, expense_id):
    if request.method != 'POST':
        return redirect('supplier_payment_overview')
    expense = get_object_or_404(SupplierExpense, id=expense_id)
    new_status = request.POST.get('status')
    if new_status in dict(SupplierExpense.STATUS_CHOICES):
        expense.status = new_status
        expense.save(update_fields=['status'])
        expense.invoice.recalculate_total()
    return redirect(request.POST.get('next', reverse('supplier_payment_detail', args=[expense.supplier_name])))


@login_required
@superuser_or_owner_required
def supplier_list(request):
    suppliers = Supplier.objects.prefetch_related('supplier_services').all()
    return render(request, 'tour_quote/supplier_list.html', {'suppliers': suppliers})


@login_required
@superuser_or_owner_required
def supplier_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Supplier name is required.')
        elif Supplier.objects.filter(name=name).exists():
            messages.error(request, f'A supplier named "{name}" already exists.')
        else:
            supplier = Supplier.objects.create(
                name=name,
                contact_person=request.POST.get('contact_person', '').strip() or None,
                email=request.POST.get('email', '').strip() or None,
                phone=request.POST.get('phone', '').strip() or None,
                address=request.POST.get('address', '').strip() or None,
                notes=request.POST.get('notes', '').strip() or None,
            )
            svc_names = request.POST.getlist('service_names')
            svc_costs = request.POST.getlist('service_costs')
            for svc_name, svc_cost in zip(svc_names, svc_costs):
                svc_name = svc_name.strip()
                if svc_name:
                    SupplierService.objects.create(
                        supplier=supplier, name=svc_name,
                        cost=safe_decimal(svc_cost),
                    )
            SupplierExpense.objects.filter(supplier_name=name, supplier__isnull=True).update(supplier=supplier)
            messages.success(request, f'Supplier "{name}" created.')
            return redirect('supplier_list')
    return render(request, 'tour_quote/supplier_form.html', {'action': 'create'})


@login_required
@superuser_or_owner_required
def supplier_edit(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Supplier name is required.')
        elif Supplier.objects.filter(name=name).exclude(pk=supplier_id).exists():
            messages.error(request, f'A supplier named "{name}" already exists.')
        else:
            old_name = supplier.name
            supplier.name = name
            supplier.contact_person = request.POST.get('contact_person', '').strip() or None
            supplier.email = request.POST.get('email', '').strip() or None
            supplier.phone = request.POST.get('phone', '').strip() or None
            supplier.address = request.POST.get('address', '').strip() or None
            supplier.notes = request.POST.get('notes', '').strip() or None
            supplier.save()
            # Rebuild services: delete all then recreate from submitted names
            supplier.supplier_services.all().delete()
            svc_names = request.POST.getlist('service_names')
            svc_costs = request.POST.getlist('service_costs')
            for svc_name, svc_cost in zip(svc_names, svc_costs):
                svc_name = svc_name.strip()
                if svc_name:
                    SupplierService.objects.create(
                        supplier=supplier, name=svc_name,
                        cost=safe_decimal(svc_cost),
                    )
            if old_name != name:
                SupplierExpense.objects.filter(supplier=supplier).update(supplier_name=name)
            messages.success(request, f'Supplier "{name}" updated.')
            return redirect('supplier_list')
    existing_services = list(supplier.supplier_services.values('name', 'cost'))
    return render(request, 'tour_quote/supplier_form.html', {
        'supplier': supplier,
        'existing_services': existing_services,
        'action': 'edit',
    })


@login_required
@superuser_or_owner_required
def supplier_delete(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    if request.method == 'POST':
        name = supplier.name
        supplier.delete()
        messages.success(request, f'Supplier "{name}" deleted.')
        return redirect('supplier_list')
    return render(request, 'tour_quote/supplier_confirm_delete.html', {'supplier': supplier})


@login_required
@superuser_or_owner_required
def supplier_payment_overview(request):
    suppliers = (
        SupplierExpense.objects
        .exclude(supplier_name__isnull=True)
        .exclude(supplier_name='')
        .values('supplier_name')
        .annotate(
            total_amount=Coalesce(Sum('amount'), Decimal('0')),
            total_count=Count('id'),
            pending_amount=Coalesce(Sum(
                Case(When(status='Pending', then=F('amount')), default=Decimal('0'), output_field=DecimalField())
            ), Decimal('0')),
            pending_count=Coalesce(Sum(
                Case(When(status='Pending', then=1), default=0, output_field=IntegerField())
            ), 0),
            paid_amount=Coalesce(Sum(
                Case(When(status='Paid', then=F('amount')), default=Decimal('0'), output_field=DecimalField())
            ), Decimal('0')),
        )
        .order_by('supplier_name')
    )
    grand_total = sum(s['total_amount'] for s in suppliers)
    grand_pending = sum(s['pending_amount'] for s in suppliers)

    paginator = Paginator(suppliers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'suppliers': page_obj,
        'supplier_count': paginator.count,
        'grand_total': grand_total,
        'grand_pending': grand_pending,
    }
    return render(request, 'tour_quote/supplier_payment_overview.html', context)


@login_required
@superuser_or_owner_required
def supplier_payment_detail(request, supplier_name):
    status_filter = request.GET.get('status', 'Pending')
    supplier_obj = Supplier.objects.prefetch_related('supplier_services').filter(name=supplier_name).first()
    expenses_qs = (
        SupplierExpense.objects
        .filter(supplier_name=supplier_name)
        .select_related('invoice', 'invoice__tour_package')
        .order_by('status', 'due_date')
    )
    if status_filter != 'all':
        expenses_qs = expenses_qs.filter(status=status_filter)

    total = sum(e.amount for e in expenses_qs)

    context = {
        'supplier_name': supplier_name,
        'supplier': supplier_obj,
        'expenses': expenses_qs,
        'total': total,
        'status_filter': status_filter,
        'status_choices': SupplierExpense.STATUS_CHOICES,
    }
    return render(request, 'tour_quote/supplier_payment_detail.html', context)


@login_required
@superuser_or_owner_required
def pending_payment_list(request):
    """List all supplier payments with filters for status and due date range."""
    from django.db.models import Q

    # Get filter parameters
    status_filter = request.GET.get('status', '')
    due_date_from = request.GET.get('due_date_from', '')
    due_date_to = request.GET.get('due_date_to', '')
    supplier_filter = request.GET.get('supplier', '')

    # Base queryset
    expenses = SupplierExpense.objects.select_related('invoice', 'supplier').all()

    # Apply filters
    if status_filter:
        expenses = expenses.filter(status=status_filter)
    if due_date_from:
        expenses = expenses.filter(due_date__gte=due_date_from)
    if due_date_to:
        expenses = expenses.filter(due_date__lte=due_date_to)
    if supplier_filter:
        expenses = expenses.filter(supplier_name__icontains=supplier_filter)

    # Order by due date (nulls last), then by supplier name
    expenses = expenses.order_by('due_date', 'supplier_name')

    # Calculate totals before pagination
    total_amount = sum(e.amount for e in expenses)
    pending_amount = sum(e.amount for e in expenses if e.status == 'Pending')
    paid_amount = sum(e.amount for e in expenses if e.status == 'Paid')
    expense_count = expenses.count()

    # Paginate
    paginator = Paginator(expenses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get unique suppliers for filter dropdown
    suppliers = SupplierExpense.objects.exclude(supplier_name__isnull=True).exclude(supplier_name='').values_list('supplier_name', flat=True).distinct().order_by('supplier_name')

    context = {
        'page_obj': page_obj,
        'expenses': page_obj,
        'total_amount': total_amount,
        'pending_amount': pending_amount,
        'paid_amount': paid_amount,
        'expense_count': expense_count,
        'status_filter': status_filter,
        'due_date_from': due_date_from,
        'due_date_to': due_date_to,
        'supplier_filter': supplier_filter,
        'suppliers': suppliers,
        'status_choices': SupplierExpense.STATUS_CHOICES,
    }
    return render(request, 'tour_quote/pending_payment_list.html', context)
